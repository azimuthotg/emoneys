from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from django.db.models import Q
from django.db import IntegrityError
import json

from .forms import LoginForm, ReceiptEditRequestForm, EditRequestApprovalForm, ReceiptEditRequestItemFormSet
from .models import Permission, Role, UserRole, Receipt, ReceiptTemplate, ReceiptItem, Department, ReceiptEditRequest, ReceiptChangeLog, User, UserActivityLog, ReceiptCancelRequest


def login_view(request):
    """
    File-based Authentication Login View
    Features:
    - Thai ID validation
    - Toast notifications
    - Approval status checking
    - Mobile API support
    """
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        # Handle both web and API requests
        content_type = request.content_type
        if content_type and 'application/json' in content_type:
            return handle_api_login(request)
        else:
            return handle_web_login(request)
    
    form = LoginForm()
    return render(request, 'accounts/login.html', {
        'title': 'เข้าสู่ระบบ',
        'form': form
    })


def handle_web_login(request):
    """Handle web-based login"""
    form = LoginForm(data=request.POST)
    
    if form.is_valid():
        ldap_uid = form.cleaned_data['username']
        password = form.cleaned_data['password']
        
        # Authenticate user
        user = authenticate(request, username=ldap_uid, password=password)
        
        if user:
            if user.is_pending_approval:
                messages.warning(
                    request, 
                    'บัญชีของคุณอยู่ระหว่างการอนุมัติ กรุณารอการติดต่อจากผู้ดูแลระบบ'
                )
            elif user.is_suspended:
                messages.error(
                    request, 
                    'บัญชีของคุณถูกระงับการใช้งาน กรุณาติดต่อผู้ดูแลระบบ'
                )
            else:
                auth_login(request, user)

                # Handle Remember Me
                remember_me = request.POST.get('remember_me')
                if remember_me:
                    # Set session expiry to 7 days (604800 seconds)
                    request.session.set_expiry(604800)
                else:
                    # Session expires when browser closes
                    request.session.set_expiry(0)

                # Log successful login
                UserActivityLog.log_login(user, request)

                messages.success(
                    request,
                    f'ยินดีต้อนรับ {user.full_name or user.username}'
                )
                return redirect('dashboard')
        else:
            # Log failed login attempt
            UserActivityLog.log_failed_login(ldap_uid, request, "รหัสบัตรประชาชนหรือรหัสผ่านไม่ถูกต้อง")

            # Check if this is a new NPU user that was just created
            try:
                from django.contrib.auth import get_user_model
                User = get_user_model()
                pending_user = User.objects.get(ldap_uid=ldap_uid, approval_status='pending')
                messages.info(
                    request,
                    f'บัญชีของคุณถูกสร้างใหม่จาก NPU แล้ว กรุณารอการอนุมัติจากผู้ดูแลระบบ'
                )
            except User.DoesNotExist:
                messages.error(request, 'รหัสบัตรประชาชนหรือรหัสผ่านไม่ถูกต้อง หรือไม่พบข้อมูลใน NPU')
    else:
        # Form validation errors
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(request, error)
    
    return render(request, 'accounts/login.html', {
        'title': 'เข้าสู่ระบบ',
        'form': form
    })


@csrf_exempt
def handle_api_login(request):
    """Handle JSON API login for mobile apps"""
    try:
        data = json.loads(request.body)
        ldap_uid = data.get('ldap_uid', '').strip()
        password = data.get('password', '')
        
        # Validate input
        if not ldap_uid or not password:
            return JsonResponse({
                'success': False,
                'message': 'กรุณากรอกรหัสบัตรประชาชนและรหัสผ่าน'
            }, status=400)
        
        # Validate Thai ID format
        if ldap_uid != 'admin' and (not ldap_uid.isdigit() or len(ldap_uid) != 13):
            return JsonResponse({
                'success': False,
                'message': 'รหัสบัตรประชาชนต้องเป็นตัวเลข 13 หลัก'
            }, status=400)
        
        # Authenticate user
        user = authenticate(request, username=ldap_uid, password=password)
        
        if user:
            if user.is_pending_approval:
                return JsonResponse({
                    'success': False,
                    'message': 'บัญชีของคุณอยู่ระหว่างการอนุมัติ'
                }, status=403)
            elif user.is_suspended:
                return JsonResponse({
                    'success': False,
                    'message': 'บัญชีของคุณถูกระงับการใช้งาน'
                }, status=403)
            else:
                auth_login(request, user)

                # Log successful login (API)
                UserActivityLog.log_login(user, request)

                return JsonResponse({
                    'success': True,
                    'message': f'ยินดีต้อนรับ {user.full_name or user.username}',
                    'user': {
                        'id': user.id,
                        'username': user.username,
                        'full_name': user.full_name,
                        'department': user.department,
                        'position_title': user.position_title,
                        'roles': [role.name for role in user.get_roles()],
                    }
                })
        else:
            # Log failed login (API)
            UserActivityLog.log_failed_login(ldap_uid, request, "API login failed: Invalid credentials")

            return JsonResponse({
                'success': False,
                'message': 'รหัสบัตรประชาชนหรือรหัสผ่านไม่ถูกต้อง'
            }, status=401)
            
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'ข้อมูลไม่ถูกต้อง'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': 'เกิดข้อผิดพลาดในระบบ'
        }, status=500)


@login_required
def dashboard(request):
    """
    User Dashboard for Receipt System
    Routes to admin dashboard for admin users
    Features:
    - User information display
    - Receipt statistics from actual Receipt model
    - Quick actions for receipt management
    - Recent receipt activities
    """
    
    # Check if user is admin - route to admin dashboard
    if request.user.is_staff or request.user.is_superuser:
        return admin_dashboard(request)
    
    # Get user's department access
    user_departments = Department.objects.none()  # Empty queryset
    if hasattr(request.user, 'department') and request.user.department:
        # Find Department objects that match user's department name
        user_departments = Department.objects.filter(name=request.user.department)
        
        # If no matching department found, create it or handle gracefully
        if not user_departments.exists():
            print(f"Warning: Department '{request.user.department}' not found in Department table")
    
    # Calculate actual receipt statistics for user's department(s)
    from django.db.models import Count, Sum
    from datetime import datetime, timedelta
    
    # Base queryset - user's department receipts
    base_queryset = Receipt.objects.all()
    if user_departments.exists():
        base_queryset = base_queryset.filter(department__in=user_departments)
    
    # Current month range
    from django.utils import timezone
    today = timezone.now().date()
    first_day_month = timezone.make_aware(
        datetime.combine(today.replace(day=1), datetime.min.time())
    )
    
    # Statistics calculations
    stats = {
        'total_receipts_issued': base_queryset.filter(created_by=request.user).count(),
        'total_receipts_approved': base_queryset.filter(status='completed').count(),
        'pending_approval': base_queryset.filter(status='draft').count(),
        'this_month_count': base_queryset.filter(
            created_at__gte=first_day_month,
            status='completed'
        ).count(),
        'this_month_amount': base_queryset.filter(
            created_at__gte=first_day_month,
            status='completed'
        ).aggregate(total=Sum('total_amount'))['total'] or 0,
        'user_total_amount': base_queryset.filter(
            created_by=request.user,
            status='completed'
        ).aggregate(total=Sum('total_amount'))['total'] or 0
    }
    
    # Recent receipts issued by current user (last 5)
    recent_received = Receipt.objects.filter(
        created_by=request.user
    ).order_by('-created_at')[:5]
    
    # Recent receipts in user's department (last 5, excluding user's own)
    recent_forwarded = Receipt.objects.filter(
        department__in=user_departments
    ).exclude(
        created_by=request.user
    ).order_by('-created_at')[:5] if user_departments.exists() else []

    # นับคำขอแก้ไขรออนุมัติ (สำหรับ Dep Manager / Senior Manager)
    pending_edit_requests = 0
    pending_cancel_requests = 0

    if request.user.has_permission('receipt_edit_approve') or request.user.has_permission('receipt_edit_approve_manager'):
        # นับคำขอแก้ไขที่รออนุมัติในแผนกตัวเอง
        edit_requests = ReceiptEditRequest.objects.filter(
            receipt__department__name=request.user.department,
            status='pending'
        )

        # กรองเฉพาะคำขอที่ผู้ใช้มีสิทธิ์อนุมัติ
        for req in edit_requests:
            if req.can_be_approved_by(request.user):
                pending_edit_requests += 1

    if request.user.has_permission('receipt_cancel_approve') or request.user.has_permission('receipt_cancel_approve_manager'):
        # นับคำขอยกเลิกที่รออนุมัติในแผนกตัวเอง
        cancel_requests = ReceiptCancelRequest.objects.filter(
            receipt__department__name=request.user.department,
            status='pending'
        )

        # กรองเฉพาะคำขอที่ผู้ใช้มีสิทธิ์อนุมัติ
        for req in cancel_requests:
            if req.can_be_approved_by(request.user):
                pending_cancel_requests += 1

    context = {
        'title': 'หน้าหลัก',
        'user': request.user,
        'stats': stats,
        'recent_received': recent_received,
        'recent_forwarded': recent_forwarded,
        'pending_edit_requests': pending_edit_requests,
        'pending_cancel_requests': pending_cancel_requests,
    }

    return render(request, 'accounts/dashboard.html', context)


@login_required
def profile_view(request):
    """User profile page with role information"""
    user_roles = request.user.get_roles()
    
    context = {
        'title': 'ข้อมูลส่วนตัว',
        'user': request.user,
        'user_roles': user_roles,
        'has_any_role': user_roles.exists(),
    }
    
    return render(request, 'accounts/profile.html', context)


@login_required
def admin_dashboard(request):
    """
    Admin Dashboard with user approval functionality
    Features:
    - Admin statistics
    - Pending user approval list
    - Quick admin actions
    """
    from django.contrib.auth import get_user_model
    User = get_user_model()
    
    # Admin statistics
    admin_stats = {
        'total_users': User.objects.count(),
        'pending_users': User.objects.filter(approval_status='pending').count(),
        'approved_users': User.objects.filter(approval_status='approved').count(),
        'total_receipts': 0,  # Placeholder for receipt count
    }
    
    # Get pending users for approval
    pending_users = User.objects.filter(approval_status='pending').order_by('-date_joined')
    
    context = {
        'title': 'Dashboard Admin',
        'user': request.user,
        'admin_stats': admin_stats,
        'pending_users': pending_users,
    }
    
    return render(request, 'accounts/admin_dashboard.html', context)


@login_required
def approve_user_ajax(request, user_id):
    """AJAX endpoint to approve user"""
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'success': False, 'message': 'ไม่มีสิทธิ์ในการดำเนินการ'}, status=403)
    
    if request.method == 'POST':
        try:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            user = User.objects.get(id=user_id)
            
            if user.approval_status == 'pending':
                user.approve_user()
                return JsonResponse({
                    'success': True, 
                    'message': f'อนุมัติผู้ใช้ {user.full_name or user.username} เรียบร้อย'
                })
            else:
                return JsonResponse({
                    'success': False, 
                    'message': f'ผู้ใช้อยู่ในสถานะ {user.approval_status} แล้ว'
                })
                
        except User.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'ไม่พบผู้ใช้'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)


@login_required
def reject_user_ajax(request, user_id):
    """AJAX endpoint to reject user"""
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'success': False, 'message': 'ไม่มีสิทธิ์ในการดำเนินการ'}, status=403)
    
    if request.method == 'POST':
        try:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            user = User.objects.get(id=user_id)
            
            if user.approval_status == 'pending':
                user.reject_user()
                return JsonResponse({
                    'success': True, 
                    'message': f'ปฏิเสธผู้ใช้ {user.full_name or user.username} เรียบร้อย'
                })
            else:
                return JsonResponse({
                    'success': False, 
                    'message': f'ผู้ใช้อยู่ในสถานะ {user.approval_status} แล้ว'
                })
                
        except User.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'ไม่พบผู้ใช้'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)


@login_required
def user_details_ajax(request, user_id):
    """AJAX endpoint to get user details"""
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'success': False, 'message': 'ไม่มีสิทธิ์ในการดำเนินการ'}, status=403)
    
    try:
        from django.contrib.auth import get_user_model
        User = get_user_model()
        user = User.objects.get(id=user_id)
        
        user_data = {
            'ldap_uid': user.ldap_uid,
            'full_name': user.full_name,
            'department': user.department,
            'position_title': user.position_title,
            'staff_type': user.staff_type,
            'employment_status': user.employment_status,
            'contact_email': user.contact_email,
            'approval_status': user.approval_status,
            'date_joined': user.date_joined.strftime('%d/%m/%Y %H:%M'),
            'last_login': user.last_login.strftime('%d/%m/%Y %H:%M') if user.last_login else 'ไม่เคยเข้าใช้',
        }
        
        html = f"""
        <div class="row g-3">
            <div class="col-md-6">
                <label class="form-label fw-bold">รหัสบัตรประชาชน:</label>
                <p class="form-control-plaintext">{user_data['ldap_uid']}</p>
            </div>
            <div class="col-md-6">
                <label class="form-label fw-bold">ชื่อ-นามสกุล:</label>
                <p class="form-control-plaintext">{user_data['full_name'] or 'ไม่ระบุ'}</p>
            </div>
            <div class="col-md-6">
                <label class="form-label fw-bold">หน่วยงาน:</label>
                <p class="form-control-plaintext">{user_data['department'] or 'ไม่ระบุ'}</p>
            </div>
            <div class="col-md-6">
                <label class="form-label fw-bold">ตำแหน่ง:</label>
                <p class="form-control-plaintext">{user_data['position_title'] or 'ไม่ระบุ'}</p>
            </div>
            <div class="col-md-6">
                <label class="form-label fw-bold">ประเภทบุคลากร:</label>
                <p class="form-control-plaintext">{user_data['staff_type'] or 'ไม่ระบุ'}</p>
            </div>
            <div class="col-md-6">
                <label class="form-label fw-bold">สถานะการทำงาน:</label>
                <p class="form-control-plaintext">{user_data['employment_status'] or 'ไม่ระบุ'}</p>
            </div>
            <div class="col-md-6">
                <label class="form-label fw-bold">อีเมล:</label>
                <p class="form-control-plaintext">{user_data['contact_email'] or 'ไม่ระบุ'}</p>
            </div>
            <div class="col-md-6">
                <label class="form-label fw-bold">สถานะการอนุมัติ:</label>
                <p class="form-control-plaintext">
                    <span class="badge bg-warning">{user_data['approval_status']}</span>
                </p>
            </div>
            <div class="col-md-6">
                <label class="form-label fw-bold">วันที่สมัคร:</label>
                <p class="form-control-plaintext">{user_data['date_joined']}</p>
            </div>
            <div class="col-md-6">
                <label class="form-label fw-bold">เข้าใช้ล่าสุด:</label>
                <p class="form-control-plaintext">{user_data['last_login']}</p>
            </div>
        </div>
        """
        
        return JsonResponse({'success': True, 'html': html})
        
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'ไม่พบผู้ใช้'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
def department_management_view(request):
    """หน้าจัดการหน่วยงาน"""
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, 'ไม่มีสิทธิ์เข้าถึงหน้านี้')
        return redirect('dashboard')
    
    from .models import Department
    from django.contrib.auth import get_user_model
    from django.db.models import Count
    
    User = get_user_model()
    
    # Get unique departments from users (from NPU AD)
    npu_departments = User.objects.exclude(
        department__isnull=True
    ).exclude(
        department__exact=''
    ).values('department').annotate(
        user_count=Count('id')
    ).order_by('department')
    
    # Get existing department codes
    existing_departments = Department.objects.all().select_related()
    dept_codes = {dept.name: dept for dept in existing_departments}
    
    # Combine NPU departments with existing codes
    departments_data = []
    for npu_dept in npu_departments:
        dept_name = npu_dept['department']
        user_count = npu_dept['user_count']
        
        # Check if we have a code for this department
        if dept_name in dept_codes:
            dept_obj = dept_codes[dept_name]
            departments_data.append({
                'id': dept_obj.id,
                'name': dept_name,
                'code': dept_obj.code,
                'is_active': dept_obj.is_active,
                'user_count': user_count,
                'has_code': True,
                'created_at': dept_obj.created_at.isoformat() if dept_obj.created_at else None,
                'address': dept_obj.address,
                'postal_code': dept_obj.postal_code,
                'phone': dept_obj.phone,
            })
        else:
            # NPU department without code yet
            departments_data.append({
                'id': None,
                'name': dept_name,
                'code': None,
                'is_active': True,
                'user_count': user_count,
                'has_code': False,
                'created_at': None,
                'address': '',
                'postal_code': '',
                'phone': '',
            })
    
    context = {
        'title': 'จัดการหน่วยงาน',
        'departments': departments_data,
        'departments_json': json.dumps(departments_data),
    }
    
    return render(request, 'accounts/department_management.html', context)


@login_required
def department_create_view(request):
    """กำหนดชื่อย่อให้หน่วยงาน NPU (AJAX)"""
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'success': False, 'message': 'ไม่มีสิทธิ์'}, status=403)
    
    if request.method == 'POST':
        try:
            from .models import Department
            from django.contrib.auth import get_user_model
            
            data = json.loads(request.body)
            code = data.get('code', '').strip().upper()
            department_name = data.get('name', '').strip()  # Changed from 'department_name' to 'name'
            is_active = data.get('is_active', True)
            
            # Address fields
            address = data.get('address', '').strip()
            postal_code = data.get('postal_code', '').strip()
            phone = data.get('phone', '').strip()
            
            # Validate input
            if not code or not department_name:
                return JsonResponse({'success': False, 'message': 'กรุณากรอกชื่อย่อและเลือกหน่วยงาน'})
            
            # Check if this NPU department exists in users
            User = get_user_model()
            if not User.objects.filter(department=department_name).exists():
                return JsonResponse({'success': False, 'message': 'ไม่พบหน่วยงานนี้ในระบบ NPU'})
            
            # Check if department already has a code
            if Department.objects.filter(name=department_name).exists():
                return JsonResponse({'success': False, 'message': 'หน่วยงานนี้มีชื่อย่อแล้ว'})
                
            # Check if code is already used
            if Department.objects.filter(code=code).exists():
                return JsonResponse({'success': False, 'message': 'ชื่อย่อนี้ถูกใช้แล้ว'})
            
            # Create department code
            department = Department.objects.create(
                name=department_name,
                code=code,
                is_active=is_active,
                address=address,
                postal_code=postal_code,
                phone=phone
            )
            
            return JsonResponse({
                'success': True,
                'message': f'กำหนดชื่อย่อ "{code}" ให้หน่วยงาน "{department_name}" เรียบร้อย',
                'department_id': department.id
            })
            
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'ข้อมูลไม่ถูกต้อง'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)


@login_required
def department_edit_view(request, department_id):
    """แก้ไขชื่อย่อหน่วยงาน (AJAX)"""
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'success': False, 'message': 'ไม่มีสิทธิ์'}, status=403)
    
    try:
        from .models import Department
        department = Department.objects.get(id=department_id)
    except Department.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'ไม่พบหน่วยงาน'}, status=404)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            code = data.get('code', '').strip().upper()
            is_active = data.get('is_active', True)
            
            # Address fields
            address = data.get('address', '').strip()
            postal_code = data.get('postal_code', '').strip()
            phone = data.get('phone', '').strip()
            
            # Validate input
            if not code:
                return JsonResponse({'success': False, 'message': 'กรุณากรอกชื่อย่อ'})
            
            # Check if code is already used by another department
            if Department.objects.filter(code=code).exclude(id=department_id).exists():
                return JsonResponse({'success': False, 'message': 'ชื่อย่อนี้ถูกใช้แล้ว'})
            
            # Update department data
            department.code = code
            department.is_active = is_active
            department.address = address
            department.postal_code = postal_code
            department.phone = phone
            department.save()
            
            return JsonResponse({
                'success': True,
                'message': f'แก้ไขชื่อย่อของ "{department.name}" เป็น "{code}" เรียบร้อย'
            })
            
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'ข้อมูลไม่ถูกต้อง'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    elif request.method == 'GET':
        # Return department data for editing
        department_data = {
            'id': department.id,
            'code': department.code,
            'name': department.name,  # Read-only from NPU AD
            'is_active': department.is_active,
            'address': department.address,
            'postal_code': department.postal_code,
            'phone': department.phone,
        }
        return JsonResponse({'success': True, 'department': department_data})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)


@login_required
def department_delete_view(request, department_id):
    """ลบหน่วยงาน (AJAX)"""
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'success': False, 'message': 'ไม่มีสิทธิ์'}, status=403)
    
    if request.method == 'POST':
        try:
            from .models import Department
            department = Department.objects.get(id=department_id)
            
            # Check if any users are assigned to this department
            from django.contrib.auth import get_user_model
            User = get_user_model()
            user_count = User.objects.filter(department_internal_name=department.name).count()
            if user_count > 0:
                return JsonResponse({
                    'success': False, 
                    'message': f'ไม่สามารถลบได้ มีบุคลากร {user_count} คนในหน่วยงานนี้'
                })
            
            department_name = department.name
            department.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'ลบหน่วยงาน "{department_name}" เรียบร้อย'
            })
            
        except Department.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'ไม่พบหน่วยงาน'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)


@login_required
def department_activate_view(request, department_id):
    """เปิดใช้งานหน่วยงาน (AJAX)"""
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'success': False, 'message': 'ไม่มีสิทธิ์'}, status=403)
    
    if request.method == 'POST':
        try:
            from .models import Department
            department = Department.objects.get(id=department_id)
            
            department.is_active = True
            department.save()
            
            return JsonResponse({
                'success': True,
                'message': f'เปิดใช้งานหน่วยงาน "{department.name}" เรียบร้อย'
            })
            
        except Department.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'ไม่พบหน่วยงาน'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)


@login_required
def department_deactivate_view(request, department_id):
    """ปิดใช้งานหน่วยงาน (AJAX)"""
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'success': False, 'message': 'ไม่มีสิทธิ์'}, status=403)
    
    if request.method == 'POST':
        try:
            from .models import Department
            department = Department.objects.get(id=department_id)
            
            department.is_active = False
            department.save()
            
            return JsonResponse({
                'success': True,
                'message': f'ปิดใช้งานหน่วยงาน "{department.name}" เรียบร้อย'
            })
            
        except Department.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'ไม่พบหน่วยงาน'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)


@login_required
def available_npu_departments_ajax(request):
    """ดึงหน่วยงาน NPU ที่ยังไม่มีชื่อย่อ (AJAX)"""
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'success': False, 'message': 'ไม่มีสิทธิ์'}, status=403)
    
    try:
        from .models import Department
        from django.contrib.auth import get_user_model
        from django.db.models import Count
        
        User = get_user_model()
        
        # Get unique departments from users (from NPU AD)
        npu_departments = User.objects.exclude(
            department__isnull=True
        ).exclude(
            department__exact=''
        ).values('department').annotate(
            user_count=Count('id')
        ).order_by('department')
        
        # Get departments that already have codes
        existing_dept_names = set(Department.objects.values_list('name', flat=True))
        
        # Filter out departments that already have codes
        available_departments = []
        for dept in npu_departments:
            if dept['department'] not in existing_dept_names:
                available_departments.append({
                    'name': dept['department'],
                    'user_count': dept['user_count']
                })
        
        return JsonResponse({
            'success': True,
            'departments': available_departments
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
def document_numbering_view(request):
    """หน้าตั้งค่าเลขที่เอกสาร"""
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, 'ไม่มีสิทธิ์เข้าถึงหน้านี้')
        return redirect('dashboard')
    
    try:
        from .models import DocumentVolume, Department
        from utils.fiscal_year import get_current_fiscal_year, format_fiscal_year_display
        from utils.notifications import get_dashboard_notifications
        from utils.fiscal_year_info import get_fiscal_year_info_card, get_volume_status_summary
        from django.db.models import Q
        
        current_fy = get_current_fiscal_year()
        
        # ดึงข้อมูลการแจ้งเตือน
        notifications = get_dashboard_notifications()
        
        # ดึงข้อมูลปีงบประมาณแบบ Real-time
        fiscal_year_info = get_fiscal_year_info_card()
        volume_summary = get_volume_status_summary(current_fy)
        
        # ดึงข้อมูลเล่มทั้งหมดในปีงบประมาณปัจจุบัน
        current_volumes = DocumentVolume.objects.filter(
            fiscal_year=current_fy
        ).select_related('department', 'created_by', 'closed_by').order_by('department__name')
        
        # ดึงข้อมูลหน่วยงานที่ยังไม่มีเล่ม
        departments_with_volumes = set(current_volumes.values_list('department_id', flat=True))
        departments_without_volumes = Department.objects.filter(
            is_active=True
        ).exclude(id__in=departments_with_volumes)
        
        # ดึงข้อมูลเล่มจากปีก่อนหน้า (สำหรับอ้างอิง)
        previous_volumes = DocumentVolume.objects.filter(
            fiscal_year=current_fy - 1
        ).select_related('department').order_by('department__name')
        
        # สถิติโดยรวม
        total_volumes = current_volumes.count()
        active_volumes = current_volumes.filter(status='active').count()
        total_documents_issued = sum(vol.last_document_number for vol in current_volumes)
        
        volumes_stats = []
        for volume in current_volumes:
            usage_percentage = volume.get_usage_percentage()
            is_nearly_full = volume.is_nearly_full()
            is_critical = usage_percentage >= 95
            
            volumes_stats.append({
                'volume': volume,
                'usage_percentage': usage_percentage,
                'is_nearly_full': is_nearly_full,
                'is_critical': is_critical,
                'remaining_capacity': volume.max_documents - volume.last_document_number
            })
        
        context = {
            'title': 'ตั้งค่าเลขที่เอกสาร',
            'current_fiscal_year': current_fy,
            'fiscal_year_display': format_fiscal_year_display(current_fy),
            'notifications': notifications,
            'fiscal_year_info': fiscal_year_info,
            'volume_summary': volume_summary,
            'volumes_stats': volumes_stats,
            'current_volumes': current_volumes,
            'departments_without_volumes': departments_without_volumes,
            'previous_volumes': previous_volumes,
            'stats': {
                'total_volumes': total_volumes,
                'active_volumes': active_volumes,
                'closed_volumes': current_volumes.filter(status='closed').count(),
                'total_documents_issued': total_documents_issued,
                'departments_count': Department.objects.filter(is_active=True).count(),
                'coverage_percentage': (total_volumes / Department.objects.filter(is_active=True).count() * 100) if Department.objects.filter(is_active=True).count() > 0 else 0
            }
        }
        
        return render(request, 'accounts/document_numbering.html', context)
        
    except ImportError as e:
        messages.error(request, f'ระบบยังไม่พร้อมใช้งาน: {str(e)}')
        return redirect('dashboard')
    except Exception as e:
        messages.error(request, f'เกิดข้อผิดพลาด: {str(e)}')
        return redirect('dashboard')




@login_required
def close_volume_ajax(request, volume_id):
    """ปิดเล่มเอกสาร (AJAX)"""
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'success': False, 'message': 'ไม่มีสิทธิ์'}, status=403)
    
    if request.method == 'POST':
        try:
            from .models import DocumentVolume, DocumentVolumeLog
            import json
            
            data = json.loads(request.body)
            reason = data.get('reason', '')
            
            try:
                volume = DocumentVolume.objects.get(id=volume_id)
            except DocumentVolume.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'ไม่พบเล่มเอกสาร'}, status=404)
            
            if volume.status == 'closed':
                return JsonResponse({'success': False, 'message': f'เล่ม {volume.volume_code} ถูกปิดแล้ว'})
            
            # ปิดเล่ม
            volume.close_volume(user=request.user)
            
            # บันทึก log
            DocumentVolumeLog.objects.create(
                volume=volume,
                action='closed',
                user=request.user,
                notes=f'ปิดเล่ม: {reason}' if reason else 'ปิดเล่มโดยผู้ดูแลระบบ'
            )
            
            return JsonResponse({
                'success': True,
                'message': f'ปิดเล่ม {volume.volume_code} เรียบร้อย',
                'volume': {
                    'id': volume.id,
                    'volume_code': volume.volume_code,
                    'status': volume.get_status_display(),
                    'closed_at': volume.closed_at.strftime('%d/%m/%Y %H:%M') if volume.closed_at else None
                }
            })
            
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'ข้อมูลไม่ถูกต้อง'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)


@login_required
def roles_permissions_view(request):
    """แสดงหน้าจัดการบทบาทและสิทธิ์"""
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, 'ไม่มีสิทธิ์เข้าถึงหน้านี้')
        return redirect('dashboard')
    
    roles = Role.objects.filter(is_active=True).prefetch_related('permissions')
    permissions = Permission.objects.filter(is_active=True)
    
    context = {
        'title': 'บทบาทและสิทธิ์',
        'roles': roles,
        'permissions': permissions,
    }
    
    return render(request, 'accounts/roles_permissions.html', context)


@login_required
def role_create_view(request):
    """สร้างบทบาทใหม่ (AJAX)"""
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'success': False, 'message': 'ไม่มีสิทธิ์'}, status=403)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            name = data.get('name', '').strip()
            display_name = data.get('display_name', '').strip()
            description = data.get('description', '').strip()
            permission_ids = data.get('permission_ids', [])
            
            # Validate input
            if not name or not display_name:
                return JsonResponse({'success': False, 'message': 'กรุณากรอกชื่อบทบาทและชื่อแสดง'})
            
            # Check if role name already exists
            if Role.objects.filter(name=name).exists():
                return JsonResponse({'success': False, 'message': 'ชื่อบทบาทนี้มีอยู่แล้ว'})
            
            # Create role
            role = Role.objects.create(
                name=name,
                display_name=display_name,
                description=description
            )
            
            # Add permissions
            if permission_ids:
                permissions = Permission.objects.filter(id__in=permission_ids, is_active=True)
                role.permissions.set(permissions)
            
            return JsonResponse({
                'success': True,
                'message': f'สร้างบทบาท "{display_name}" เรียบร้อย',
                'role_id': role.id
            })
            
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'ข้อมูลไม่ถูกต้อง'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)


@login_required
def role_edit_view(request, role_id):
    """แก้ไขบทบาท (AJAX)"""
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'success': False, 'message': 'ไม่มีสิทธิ์'}, status=403)
    
    try:
        role = Role.objects.get(id=role_id, is_active=True)
    except Role.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'ไม่พบบทบาท'}, status=404)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            display_name = data.get('display_name', '').strip()
            description = data.get('description', '').strip()
            permission_ids = data.get('permission_ids', [])
            
            # Validate input
            if not display_name:
                return JsonResponse({'success': False, 'message': 'กรุณากรอกชื่อแสดง'})
            
            # Update role
            role.display_name = display_name
            role.description = description
            role.save()
            
            # Update permissions
            if permission_ids:
                permissions = Permission.objects.filter(id__in=permission_ids, is_active=True)
                role.permissions.set(permissions)
            else:
                role.permissions.clear()
            
            return JsonResponse({
                'success': True,
                'message': f'แก้ไขบทบาท "{display_name}" เรียบร้อย'
            })
            
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'ข้อมูลไม่ถูกต้อง'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    elif request.method == 'GET':
        # Return role data for editing
        role_data = {
            'id': role.id,
            'name': role.name,
            'display_name': role.display_name,
            'description': role.description,
            'permission_ids': list(role.permissions.values_list('id', flat=True))
        }
        return JsonResponse({'success': True, 'role': role_data})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)


@login_required
def role_delete_view(request, role_id):
    """ลบบทบาท (AJAX)"""
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'success': False, 'message': 'ไม่มีสิทธิ์'}, status=403)
    
    if request.method == 'POST':
        try:
            role = Role.objects.get(id=role_id, is_active=True)
            
            # Check if any users have this role
            user_count = UserRole.objects.filter(role=role, is_active=True).count()
            if user_count > 0:
                return JsonResponse({
                    'success': False, 
                    'message': f'ไม่สามารถลบได้ มีผู้ใช้ {user_count} คนที่มีบทบาทนี้'
                })
            
            role_name = role.display_name
            role.is_active = False
            role.save()
            
            return JsonResponse({
                'success': True,
                'message': f'ลบบทบาท "{role_name}" เรียบร้อย'
            })
            
        except Role.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'ไม่พบบทบาท'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)


@login_required
def user_role_assign_view(request, user_id):
    """กำหนดบทบาทให้ผู้ใช้ (AJAX)"""
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'success': False, 'message': 'ไม่มีสิทธิ์'}, status=403)
    
    try:
        from django.contrib.auth import get_user_model
        User = get_user_model()
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'ไม่พบผู้ใช้'}, status=404)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            role_ids = data.get('role_ids', [])
            
            # Clear existing roles
            UserRole.objects.filter(user=user).update(is_active=False)
            
            # Assign new roles
            if role_ids:
                roles = Role.objects.filter(id__in=role_ids, is_active=True)
                for role in roles:
                    user.assign_role(role, assigned_by=request.user)
            
            return JsonResponse({
                'success': True,
                'message': f'กำหนดบทบาทให้ {user.get_display_name()} เรียบร้อย'
            })
            
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'ข้อมูลไม่ถูกต้อง'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    elif request.method == 'GET':
        # Return user roles for editing
        user_roles = user.get_roles()
        user_data = {
            'id': user.id,
            'name': user.get_display_name(),
            'role_ids': list(user_roles.values_list('id', flat=True))
        }
        return JsonResponse({'success': True, 'user': user_data})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)


@login_required
def user_management_view(request):
    """
    User Management Page - separate from dashboard
    Features:
    - Tabbed interface for different user statuses
    - Bulk operations
    - Role assignment
    - Search and filter functionality
    """
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, 'ไม่มีสิทธิ์เข้าถึงหน้านี้')
        return redirect('dashboard')
    
    from django.contrib.auth import get_user_model
    User = get_user_model()
    
    # Get user statistics
    stats = {
        'total_users': User.objects.count(),
        'pending_users': User.objects.filter(approval_status='pending').count(),
        'approved_users': User.objects.filter(approval_status='approved', is_active=True).count(),  # Only active approved users
        'suspended_users': User.objects.filter(is_active=False).count(),
    }
    
    # Get users for each tab
    pending_users = User.objects.filter(approval_status='pending').order_by('-date_joined')
    approved_users = User.objects.filter(approval_status='approved', is_active=True).order_by('-approved_at')
    suspended_users = User.objects.filter(is_active=False).order_by('-date_joined')
    all_users = User.objects.all().order_by('-date_joined')
    
    # Get available roles for filter dropdown
    available_roles = Role.objects.filter(is_active=True)
    
    context = {
        'title': 'จัดการผู้ใช้งาน',
        'stats': stats,
        'pending_users': pending_users,
        'approved_users': approved_users,
        'suspended_users': suspended_users,
        'all_users': all_users,
        'available_roles': available_roles,
    }
    
    return render(request, 'accounts/user_management.html', context)


@login_required
def suspend_user_ajax(request, user_id):
    """AJAX endpoint to suspend user"""
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'success': False, 'message': 'ไม่มีสิทธิ์ในการดำเนินการ'}, status=403)
    
    if request.method == 'POST':
        try:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            user = User.objects.get(id=user_id)
            
            if user.is_active:
                user.is_active = False
                user.save()
                return JsonResponse({
                    'success': True, 
                    'message': f'ระงับผู้ใช้ {user.full_name or user.username} เรียบร้อย'
                })
            else:
                return JsonResponse({
                    'success': False, 
                    'message': 'ผู้ใช้ถูกระงับอยู่แล้ว'
                })
                
        except User.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'ไม่พบผู้ใช้'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)


@login_required
def activate_user_ajax(request, user_id):
    """AJAX endpoint to activate user"""
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'success': False, 'message': 'ไม่มีสิทธิ์ในการดำเนินการ'}, status=403)
    
    if request.method == 'POST':
        try:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            user = User.objects.get(id=user_id)
            
            if not user.is_active:
                user.is_active = True
                user.save()
                return JsonResponse({
                    'success': True, 
                    'message': f'เปิดใช้งานผู้ใช้ {user.full_name or user.username} เรียบร้อย'
                })
            else:
                return JsonResponse({
                    'success': False, 
                    'message': 'ผู้ใช้ใช้งานได้อยู่แล้ว'
                })
                
        except User.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'ไม่พบผู้ใช้'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)


@login_required
def get_available_roles_ajax(request):
    """AJAX endpoint to get available roles"""
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'success': False, 'message': 'ไม่มีสิทธิ์ในการดำเนินการ'}, status=403)
    
    if request.method == 'GET':
        roles = Role.objects.filter(is_active=True).values('id', 'name', 'display_name', 'description')
        return JsonResponse({'success': True, 'roles': list(roles)})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)


@login_required
def user_roles_ajax(request, user_id):
    """AJAX endpoint to get/set user roles"""
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'success': False, 'message': 'ไม่มีสิทธิ์ในการดำเนินการ'}, status=403)
    
    try:
        from django.contrib.auth import get_user_model
        User = get_user_model()
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'ไม่พบผู้ใช้'}, status=404)
    
    if request.method == 'GET':
        # Return user data for role assignment
        user_roles = user.get_roles()
        user_data = {
            'id': user.id,
            'full_name': user.full_name or user.username,
            'ldap_uid': user.ldap_uid,
            'department': user.department or 'ไม่ระบุ',
            'role_ids': list(user_roles.values_list('id', flat=True)),  # Get role IDs for matching
        }
        return JsonResponse({'success': True, 'user': user_data})
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            role_ids = data.get('role_ids', [])
            
            # Clear existing roles
            UserRole.objects.filter(user=user).update(is_active=False)
            
            # Assign new roles based on role IDs from database
            if role_ids:
                roles = Role.objects.filter(id__in=role_ids, is_active=True)
                for role in roles:
                    user.assign_role(role, assigned_by=request.user)
            
            return JsonResponse({
                'success': True,
                'message': f'อัปเดตบทบาทผู้ใช้ {user.full_name or user.username} เรียบร้อย'
            })
            
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'ข้อมูลไม่ถูกต้อง'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)


def logout_view(request):
    """Logout user and redirect to login"""
    if request.user.is_authenticated:
        # Log logout before actually logging out
        UserActivityLog.log_logout(request.user, request)

        messages.success(request, f'ออกจากระบบเรียบร้อย {request.user.full_name or request.user.username}')

    auth_logout(request)
    return redirect('login')


def receipt_verify_view(request, verification_hash=None):
    """
    หน้าตรวจสอบใบสำคัญรับเงินผ่าน QR Code
    """
    context = {
        'title': 'ตรวจสอบใบสำคัญรับเงิน',
        'receipt': None,
        'found': True
    }
    
    # ถ้ามี hash จาก URL parameter
    if verification_hash:
        try:
            receipt = Receipt.objects.select_related('department', 'created_by').prefetch_related('items').get(
                verification_hash=verification_hash
            )
            
            # ตรวจสอบความถูกต้องของ hash
            if receipt.verify_integrity():
                context['receipt'] = receipt
            else:
                context['found'] = False
                messages.error(request, 'ข้อมูลใบสำคัญรับเงินไม่ตรงกับที่บันทึกในระบบ')
                
        except Receipt.DoesNotExist:
            context['found'] = False
            messages.error(request, 'ไม่พบใบสำคัญรับเงินที่ต้องการตรวจสอบ')
    
    # ถ้ามี hash จาก GET parameter  
    elif request.GET.get('hash'):
        hash_param = request.GET.get('hash').strip()
        if len(hash_param) == 64:  # SHA-256 hash length
            return redirect('receipt_verify', verification_hash=hash_param)
        else:
            messages.error(request, 'รหัสตรวจสอบไม่ถูกต้อง (ต้องเป็น 64 ตัวอักษร)')
    
    return render(request, 'accounts/receipt_verify.html', context)


@login_required
def receipt_qr_image_view(request, receipt_id):
    """
    สร้างภาพ QR Code สำหรับใบสำคัญรับเงิน
    """
    try:
        receipt = Receipt.objects.get(id=receipt_id)

        # Import QR generator
        from utils.qr_generator import create_verification_qr_for_receipt

        # บันทึก receipt ก่อนสร้าง QR (กรณียังไม่มี hash)
        if not receipt.verification_hash or not receipt.qr_code_data:
            receipt.save()

        # สร้าง QR Code เป็น base64
        qr_base64 = create_verification_qr_for_receipt(receipt)

        return JsonResponse({
            'success': True,
            'qr_image': qr_base64,
            'verification_url': receipt.get_verification_url(),
            'hash': receipt.verification_hash
        })

    except Receipt.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'ไม่พบใบสำคัญรับเงิน'}, status=404)
    except Exception as e:
        import traceback
        print(f"QR Generation Error: {traceback.format_exc()}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
def receipt_create_view(request):
    """
    หน้าสร้างใบสำคัญรับเงิน Frontend
    """
    # ตรวจสอบสิทธิ์การสร้างใบสำคัญ
    if not request.user.has_permission('receipt_create'):
        messages.error(request, 'คุณไม่มีสิทธิ์สร้างใบสำคัญรับเงิน')
        return redirect('dashboard')
    
    # ดึงหน่วยงานของผู้ใช้
    try:
        user_department = Department.objects.get(name=request.user.department)
    except Department.DoesNotExist:
        messages.error(request, 'ไม่พบหน่วยงานของคุณในระบบ กรุณาติดต่อผู้ดูแล')
        return redirect('dashboard')
    
    # ดึงรายการสำเร็จรูป
    receipt_templates = ReceiptTemplate.objects.filter(is_active=True).order_by('category', 'name')
    
    context = {
        'title': 'สร้างใบสำคัญรับเงิน',
        'department': user_department,
        'receipt_templates': receipt_templates,
        'user_full_name': request.user.get_display_name(),
    }
    
    return render(request, 'accounts/receipt_create.html', context)


@login_required  
def receipt_list_view(request):
    """
    รายการใบสำคัญรับเงินของผู้ใช้
    """
    # ตรวจสอบสิทธิ์
    if not request.user.has_permission('receipt_view_own'):
        messages.error(request, 'คุณไม่มีสิทธิ์ดูรายการใบสำคัญรับเงิน')
        return redirect('dashboard')
    
    # Filter receipts based on user permissions
    if request.user.has_permission('receipt_view_all'):
        # ดูได้ทั้งหมด
        receipts = Receipt.objects.all()
        view_scope = 'ทั้งหมด'
    elif request.user.has_permission('receipt_view_department'):
        # ดูได้เฉพาะหน่วยงาน
        receipts = Receipt.objects.filter(department__name=request.user.department)
        view_scope = f'หน่วยงาน: {request.user.department}'
    elif request.user.has_permission('receipt_view_own'):
        # Basic Users can see all receipts in their department for receipt number tracking
        receipts = Receipt.objects.filter(department__name=request.user.department)
        view_scope = f'หน่วยงาน: {request.user.department}'
    else:
        # ไม่มีสิทธิ์ดูใด ๆ
        receipts = Receipt.objects.none()
        view_scope = 'ไม่มีสิทธิ์'
    
    # Apply filters
    status_filter = request.GET.get('status')
    if status_filter:
        receipts = receipts.filter(status=status_filter)
    
    search_query = request.GET.get('q')
    if search_query:
        receipts = receipts.filter(
            Q(receipt_number__icontains=search_query) |
            Q(recipient_name__icontains=search_query)
        )
    
    # Order and paginate
    receipts = receipts.select_related('department', 'created_by').order_by('-created_at')
    
    paginator = Paginator(receipts, 20)
    page = request.GET.get('page')
    receipts_page = paginator.get_page(page)
    
    # Check if user can request edit
    user_can_request_edit = request.user.has_permission('receipt_edit_request')
    
    # Check if user can approve edit requests
    user_can_approve_edit = (
        request.user.has_permission('receipt_edit_approve') or 
        request.user.has_permission('receipt_edit_approve_manager') or
        request.user.has_permission('receipt_view_all')
    )
    
    # Add edit request status and cancel request status to each receipt
    for receipt in receipts_page:
        pending_request = receipt.edit_requests.filter(status='pending').first()
        approved_request = receipt.edit_requests.filter(status='approved').first()
        applied_request = receipt.edit_requests.filter(status='applied').first()
        
        if pending_request:
            receipt.edit_status = 'pending'
            receipt.pending_request = pending_request
            # Check if current user can approve this specific request
            receipt.can_approve_pending = pending_request.can_be_approved_by(request.user)
        elif applied_request:
            receipt.edit_status = 'applied'
            receipt.applied_request = applied_request
        elif approved_request:
            receipt.edit_status = 'approved'
            receipt.approved_request = approved_request
        else:
            receipt.edit_status = None
        
        # Add cancel request status
        pending_cancel_request = receipt.cancel_requests.filter(status='pending').first()
        rejected_cancel_request = receipt.cancel_requests.filter(status='rejected').first()
        
        if pending_cancel_request:
            receipt.cancel_status = 'pending'
            receipt.pending_cancel_request = pending_cancel_request
            # Check if current user can approve this cancel request
            receipt.can_approve_cancel_request = pending_cancel_request.can_be_approved_by(request.user)
        elif rejected_cancel_request:
            receipt.cancel_status = 'rejected'
            receipt.rejected_cancel_request = rejected_cancel_request
        else:
            receipt.cancel_status = None
    
    context = {
        'title': 'รายการใบสำคัญรับเงิน',
        'receipts': receipts_page,
        'view_scope': view_scope,
        'status_filter': status_filter,
        'search_query': search_query,
        'status_choices': Receipt.STATUS_CHOICES,
        'user_can_request_edit': user_can_request_edit,
        'user_can_approve_edit': user_can_approve_edit,
    }
    
    return render(request, 'accounts/receipt_list.html', context)


@login_required
def receipt_detail_view(request, receipt_id):
    """
    รายละเอียดใบสำคัญรับเงิน
    """
    try:
        receipt = Receipt.objects.select_related('department', 'created_by').prefetch_related('items').get(id=receipt_id)
        
        # ตรวจสอบสิทธิ์ในการดู
        can_view = False
        
        if request.user.has_permission('receipt_view_all'):
            can_view = True
        elif request.user.has_permission('receipt_view_department') and receipt.department.name == request.user.department:
            can_view = True
        elif request.user.has_permission('receipt_view_own'):
            # Basic Users can view all receipts in their department for receipt number tracking
            can_view = (receipt.department.name == request.user.department)
        else:
            can_view = False
        
        if not can_view:
            messages.error(request, 'คุณไม่มีสิทธิ์ดูใบสำคัญรับเงินนี้')
            return redirect('receipt_list')
        
        # Check for cancel request status
        pending_cancel_request = receipt.cancel_requests.filter(status='pending').first()
        rejected_cancel_request = receipt.cancel_requests.filter(status='rejected').first()

        # Check for edit request status
        pending_edit_request = receipt.edit_requests.filter(status='pending').first()
        rejected_edit_request = receipt.edit_requests.filter(status='rejected').first()

        # Get change logs (audit trail)
        change_logs = receipt.change_logs.select_related('user', 'edit_request').all()

        context = {
            'title': f'ใบสำคัญรับเงิน {receipt.receipt_number}',
            'receipt': receipt,
            'can_be_cancelled_directly': receipt.can_be_cancelled_directly(request.user),
            'can_be_cancelled_by': receipt.can_be_cancelled_by(request.user),
            'pending_cancel_request': pending_cancel_request,
            'rejected_cancel_request': rejected_cancel_request,
            'pending_edit_request': pending_edit_request,
            'rejected_edit_request': rejected_edit_request,
            'change_logs': change_logs,
            'user_can_request_edit': (receipt.created_by == request.user),  # เฉพาะเจ้าของเท่านั้น
        }
        
        return render(request, 'accounts/receipt_detail.html', context)
        
    except Receipt.DoesNotExist:
        messages.error(request, 'ไม่พบใบสำคัญรับเงินที่ต้องการ')
        return redirect('receipt_list')


@login_required
def receipt_save_ajax(request):
    """
    บันทึกใบสำคัญรับเงินผ่าน AJAX
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)
    
    # ตรวจสอบสิทธิ์
    if not request.user.has_permission('receipt_create'):
        return JsonResponse({'success': False, 'message': 'ไม่มีสิทธิ์สร้างใบสำคัญรับเงิน'}, status=403)
    
    try:
        data = json.loads(request.body)
        
        # ตรวจสอบข้อมูลจำเป็น
        required_fields = ['recipient_name', 'recipient_address', 'recipient_id_card']
        for field in required_fields:
            if not data.get(field):
                return JsonResponse({'success': False, 'message': f'กรุณากรอก{field}'}, status=400)
        
        # ตรวจสอบจำนวนเงิน
        try:
            total_amount = float(data.get('total_amount', 0))
            if total_amount <= 0:
                return JsonResponse({'success': False, 'message': 'จำนวนเงินต้องมากกว่า 0'}, status=400)
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'message': 'จำนวนเงินไม่ถูกต้อง'}, status=400)
        
        # ดึงหน่วยงาน
        if not request.user.department:
            return JsonResponse({'success': False, 'message': 'ไม่พบข้อมูลหน่วยงานของคุณ'}, status=400)
            
        try:
            department = Department.objects.get(name=request.user.department)
        except Department.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'ไม่พบหน่วยงานของคุณในระบบ'}, status=400)
        
        # สร้างใบสำคัญรับเงิน
        from datetime import datetime
        from utils.fiscal_year import get_current_fiscal_year, get_fiscal_year_dates

        status = data.get('status', 'draft')

        # ตั้งค่าและตรวจสอบ receipt_date
        receipt_date = None
        if data.get('receipt_date'):
            try:
                # Parse วันที่จาก string (YYYY-MM-DD)
                receipt_date = datetime.strptime(data.get('receipt_date'), '%Y-%m-%d').date()

                # ตรวจสอบวันที่
                today = datetime.now().date()

                # ห้ามวันที่อนาคต
                if receipt_date > today:
                    return JsonResponse({'success': False, 'message': 'ไม่สามารถกำหนดวันที่ล่วงหน้าได้'}, status=400)

                # ตรวจสอบว่าไม่เกินต้นปีงบประมาณ
                current_fy = get_current_fiscal_year()
                fiscal_start, fiscal_end = get_fiscal_year_dates(current_fy)

                if receipt_date < fiscal_start:
                    return JsonResponse({
                        'success': False,
                        'message': f'วันที่ย้อนหลังเกินกำหนด (ย้อนหลังได้ถึง {fiscal_start.strftime("%d/%m/%Y")} เท่านั้น)'
                    }, status=400)

            except ValueError:
                return JsonResponse({'success': False, 'message': 'รูปแบบวันที่ไม่ถูกต้อง'}, status=400)
        elif status == 'completed':
            # ถ้าบันทึกเป็นเสร็จสิ้นแต่ไม่มีวันที่ ให้ใช้วันที่ปัจจุบัน
            receipt_date = datetime.now().date()

        receipt = Receipt.objects.create(
            department=department,
            created_by=request.user,
            recipient_name=data.get('recipient_name', ''),
            recipient_address=data.get('recipient_address', ''),
            recipient_postal_code=data.get('recipient_postal_code', ''),
            recipient_id_card=data.get('recipient_id_card', ''),
            is_loan=data.get('is_loan', False),
            total_amount=total_amount,
            status=status,
            receipt_date=receipt_date
        )
        
        # สร้างรายการ
        items_data = data.get('items', [])
        calculated_total = 0
        
        for idx, item_data in enumerate(items_data, 1):
            try:
                amount = float(item_data.get('amount', 0))
                if amount <= 0:
                    continue
                calculated_total += amount
                
                # จัดการ template_id
                template_id = item_data.get('template_id')
                if template_id and str(template_id).lower() in ['null', 'none', '']:
                    template_id = None
                elif template_id:
                    try:
                        template_id = int(template_id)
                    except (ValueError, TypeError):
                        template_id = None
                
                ReceiptItem.objects.create(
                    receipt=receipt,
                    template_id=template_id,
                    description=item_data.get('description', ''),
                    amount=amount,
                    order=idx
                )
            except (ValueError, TypeError) as e:
                return JsonResponse({'success': False, 'message': f'รายการที่ {idx} มีข้อผิดพลาด: {str(e)}'}, status=400)
        
        # อัปเดต total_amount
        receipt.total_amount = calculated_total
        receipt.save()  # จะ auto-generate ทั้ง hash และ qr code

        # บันทึก Change Log
        ReceiptChangeLog.log_change(
            receipt=receipt,
            action='created',
            user=request.user,
            notes=f'สร้างใบสำคัญรับเงิน (สถานะ: {receipt.get_status_display()})'
        )

        return JsonResponse({
            'success': True,
            'message': 'บันทึกใบสำคัญรับเงินเรียบร้อย',
            'receipt_id': receipt.id,
            'receipt_number': receipt.receipt_number,
            'receipt_date': receipt.receipt_date.strftime('%d/%m/%Y') if receipt.receipt_date else '',
            'verification_url': receipt.get_verification_url()
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'ข้อมูล JSON ไม่ถูกต้อง'}, status=400)
    except IntegrityError as e:
        import logging
        logging.error(f'Receipt save IntegrityError: {str(e)}')
        # ถ้าเกิด duplicate receipt number ให้ลองใหม่
        if 'Duplicate entry' in str(e) or 'UNIQUE constraint' in str(e):
            return JsonResponse({'success': False, 'message': 'เลขที่ใบสำคัญซ้ำ กรุณาลองใหม่อีกครั้ง'}, status=409)
        return JsonResponse({'success': False, 'message': f'เกิดข้อผิดพลาดในการบันทึก: {str(e)}'}, status=500)
    except Exception as e:
        import logging
        import traceback
        logging.error(f'Receipt save error: {str(e)}')
        logging.error(traceback.format_exc())
        return JsonResponse({'success': False, 'message': f'เกิดข้อผิดพลาด: {str(e)}'}, status=500)


@login_required
def receipt_pdf_view(request, receipt_id):
    """
    สร้างและแสดง PDF ใบสำคัญรับเงินแบบ inline
    """
    try:
        receipt = Receipt.objects.get(id=receipt_id)
        
        # ตรวจสอบสิทธิ์
        if not request.user.has_permission('receipt_view_own') and not request.user.has_permission('receipt_view_all'):
            if not request.user.has_permission('receipt_view_department'):
                messages.error(request, 'ไม่มีสิทธิ์ดูใบสำคัญรับเงิน')
                return redirect('receipt_list')
        
        # ตรวจสอบสิทธิ์เฉพาะ
        if request.user.has_permission('receipt_view_own') and receipt.created_by != request.user:
            # ถ้ามีสิทธิ์ดูเฉพาะของตัวเอง แต่ไม่ใช่เจ้าของ
            if not request.user.has_permission('receipt_view_department') and not request.user.has_permission('receipt_view_all'):
                messages.error(request, 'ไม่มีสิทธิ์ดูใบสำคัญรับเงินนี้')
                return redirect('receipt_list')
        
        if request.user.has_permission('receipt_view_department') and receipt.department.name != request.user.department:
            # ถ้ามีสิทธิ์ดูเฉพาะหน่วยงาน แต่ไม่ใช่หน่วยงานเดียวกัน
            if not request.user.has_permission('receipt_view_all'):
                messages.error(request, 'ไม่มีสิทธิ์ดูใบสำคัญรับเงินของหน่วยงานอื่น')
                return redirect('receipt_list')
        
        # สร้าง PDF แบบ inline
        from .pdf_generator import generate_receipt_pdf
        return generate_receipt_pdf(receipt, inline=True)
        
    except Receipt.DoesNotExist:
        messages.error(request, 'ไม่พบใบสำคัญรับเงินที่ต้องการ')
        return redirect('receipt_list')
    except Exception as e:
        import logging
        logging.error(f'PDF generation error: {str(e)}')
        messages.error(request, f'เกิดข้อผิดพลาดในการสร้าง PDF: {str(e)}')
        return redirect('receipt_list')


@login_required
def receipt_pdf_download_view(request, receipt_id):
    """
    ดาวน์โหลด PDF ใบสำคัญรับเงิน
    """
    try:
        receipt = Receipt.objects.get(id=receipt_id)
        
        # ตรวจสอบสิทธิ์ (เหมือนกับ receipt_pdf_view)
        if not request.user.has_permission('receipt_view_own') and not request.user.has_permission('receipt_view_all'):
            if not request.user.has_permission('receipt_view_department'):
                messages.error(request, 'ไม่มีสิทธิ์ดูใบสำคัญรับเงิน')
                return redirect('receipt_list')
        
        # ตรวจสอบสิทธิ์เฉพาะ
        if request.user.has_permission('receipt_view_own') and receipt.created_by != request.user:
            if not request.user.has_permission('receipt_view_department') and not request.user.has_permission('receipt_view_all'):
                messages.error(request, 'ไม่มีสิทธิ์ดูใบสำคัญรับเงินนี้')
                return redirect('receipt_list')
        
        if request.user.has_permission('receipt_view_department') and receipt.department.name != request.user.department:
            if not request.user.has_permission('receipt_view_all'):
                messages.error(request, 'ไม่มีสิทธิ์ดูใบสำคัญรับเงินของหน่วยงานอื่น')
                return redirect('receipt_list')
        
        # สร้าง PDF แบบ download
        from .pdf_generator import generate_receipt_pdf
        return generate_receipt_pdf(receipt, inline=False)
        
    except Receipt.DoesNotExist:
        messages.error(request, 'ไม่พบใบสำคัญรับเงินที่ต้องการ')
        return redirect('receipt_list')
    except Exception as e:
        import logging
        logging.error(f'PDF download error: {str(e)}')
        messages.error(request, f'เกิดข้อผิดพลาดในการดาวน์โหลด PDF: {str(e)}')
        return redirect('receipt_list')


@login_required
def receipt_pdf_v2_view(request, receipt_id):
    """
    ดู PDF ใบสำคัญรับเงิน เวอร์ชัน 2 (wkhtmltopdf)
    """
    try:
        import pdfkit
        from django.template.loader import render_to_string
        
        receipt = Receipt.objects.get(id=receipt_id)
        
        # ตรวจสอบสิทธิ์ (เหมือน v1)
        if not request.user.has_permission('receipt_view_own') and not request.user.has_permission('receipt_view_all'):
            if not request.user.has_permission('receipt_view_department'):
                messages.error(request, 'ไม่มีสิทธิ์ดูใบสำคัญรับเงิน')
                return redirect('receipt_list')
        
        # ตรวจสอบสิทธิ์เฉพาะ
        if request.user.has_permission('receipt_view_own') and receipt.created_by != request.user:
            # ถ้ามีสิทธิ์ดูเฉพาะของตัวเอง แต่ไม่ใช่เจ้าของ
            if not request.user.has_permission('receipt_view_department') and not request.user.has_permission('receipt_view_all'):
                messages.error(request, 'ไม่มีสิทธิ์ดูใบสำคัญรับเงินนี้')
                return redirect('receipt_list')
        
        if request.user.has_permission('receipt_view_department') and receipt.department.name != request.user.department:
            # ถ้ามีสิทธิ์ดูเฉพาะหน่วยงาน แต่ไม่ใช่หน่วยงานเดียวกัน
            if not request.user.has_permission('receipt_view_all'):
                messages.error(request, 'ไม่มีสิทธิ์ดูใบสำคัญรับเงินของหน่วยงานอื่น')
                return redirect('receipt_list')
        
        # เตรียมข้อมูลสำหรับ template
        context = {
            'receipt': receipt,
            'items': receipt.items.all().order_by('order'),
            'qr_data': f"{receipt.receipt_number}",
        }
        
        # Render HTML template
        html_content = render_to_string('accounts/receipt_pdf_v2.html', context, request)
        
        # wkhtmltopdf configuration และ options
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        
        options = {
            'page-size': 'A4',
            'encoding': 'UTF-8',
            'margin-top': '0.5in',
            'margin-right': '0.5in',
            'margin-bottom': '0.5in',
            'margin-left': '0.5in',
            'no-outline': None,
            'enable-local-file-access': None,
            'print-media-type': None,
        }
        
        # สร้าง PDF
        pdf = pdfkit.from_string(html_content, False, options=options, configuration=config)

        # ส่งกลับเป็น HTTP Response
        filename_number = receipt.receipt_number if receipt.receipt_number else f'draft_{receipt.id}'
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="receipt_{filename_number}_v2.pdf"'

        return response
        
    except Receipt.DoesNotExist:
        messages.error(request, 'ไม่พบใบสำคัญรับเงินที่ต้องการ')
        return redirect('receipt_list')
    except Exception as e:
        import logging
        logging.error(f'PDF v2 generation error: {str(e)}')
        messages.error(request, f'เกิดข้อผิดพลาดในการสร้าง PDF v2: {str(e)}')
        return redirect('receipt_list')


@login_required 
def receipt_pdf_v2_download_view(request, receipt_id):
    """
    ดาวน์โหลด PDF ใบสำคัญรับเงิน เวอร์ชัน 2 (wkhtmltopdf)
    """
    try:
        import pdfkit
        from django.template.loader import render_to_string
        
        receipt = Receipt.objects.get(id=receipt_id)
        
        # ตรวจสอบสิทธิ์ (เหมือน v1)
        if not request.user.has_permission('receipt_view_own') and not request.user.has_permission('receipt_view_all'):
            if not request.user.has_permission('receipt_view_department'):
                messages.error(request, 'ไม่มีสิทธิ์ดูใบสำคัญรับเงิน')
                return redirect('receipt_list')
        
        # ตรวจสอบสิทธิ์เฉพาะ
        if request.user.has_permission('receipt_view_own') and receipt.created_by != request.user:
            # ถ้ามีสิทธิ์ดูเฉพาะของตัวเอง แต่ไม่ใช่เจ้าของ
            if not request.user.has_permission('receipt_view_department') and not request.user.has_permission('receipt_view_all'):
                messages.error(request, 'ไม่มีสิทธิ์ดูใบสำคัญรับเงินนี้')
                return redirect('receipt_list')
        
        if request.user.has_permission('receipt_view_department') and receipt.department.name != request.user.department:
            # ถ้ามีสิทธิ์ดูเฉพาะหน่วยงาน แต่ไม่ใช่หน่วยงานเดียวกัน
            if not request.user.has_permission('receipt_view_all'):
                messages.error(request, 'ไม่มีสิทธิ์ดูใบสำคัญรับเงินของหน่วยงานอื่น')
                return redirect('receipt_list')
        
        # เตรียมข้อมูลสำหรับ template
        context = {
            'receipt': receipt,
            'items': receipt.items.all().order_by('order'),
            'qr_data': f"{receipt.receipt_number}",
        }
        
        # Render HTML template
        html_content = render_to_string('accounts/receipt_pdf_v2.html', context, request)
        
        # wkhtmltopdf configuration และ options
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        
        options = {
            'page-size': 'A4',
            'encoding': 'UTF-8',
            'margin-top': '0.5in',
            'margin-right': '0.5in',
            'margin-bottom': '0.5in',
            'margin-left': '0.5in',
            'no-outline': None,
            'enable-local-file-access': None,
            'print-media-type': None,
        }
        
        # สร้าง PDF
        pdf = pdfkit.from_string(html_content, False, options=options, configuration=config)

        # ส่งกลับเป็น HTTP Response (download)
        filename_number = receipt.receipt_number if receipt.receipt_number else f'draft_{receipt.id}'
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="receipt_{filename_number}_v2.pdf"'

        return response
        
    except Receipt.DoesNotExist:
        messages.error(request, 'ไม่พบใบสำคัญรับเงินที่ต้องการ')
        return redirect('receipt_list')
    except Exception as e:
        import logging
        logging.error(f'PDF v2 download error: {str(e)}')
        messages.error(request, f'เกิดข้อผิดพลาดในการดาวน์โหลด PDF v2: {str(e)}')
        return redirect('receipt_list')


def receipt_check_public_view(request, dept_code=None, date_part=None, number_part=None, receipt_number=None):
    """
    หน้าตรวจสอบใบสำคัญรับเงินแบบง่าย (Public - ไม่ต้อง Login)
    URL Pattern:
    - /check/ARC/260925/0005 -> dept_code=ARC, date_part=260925, number_part=0005 (ใหม่)
    - /check/260925/0005 -> date_part=260925, number_part=0005 (เก่า - backward compatibility)
    """
    # รวม date_part และ number_part เป็นเลขที่เอกสาร
    if date_part and number_part:
        receipt_number = f"{date_part}/{number_part}"

    context = {
        'title': 'ตรวจสอบใบสำคัญรับเงิน',
        'receipt': None,
        'found': False,
        'receipt_number_search': receipt_number,
        'dept_code_search': dept_code
    }

    try:
        # สร้าง query filter
        filters = {
            'receipt_number': receipt_number,
            'status': 'completed'  # แสดงเฉพาะที่เสร็จสิ้น
        }

        # ถ้ามี dept_code ให้ค้นหาเฉพาะหน่วยงานนั้น
        if dept_code:
            filters['department__code'] = dept_code

        # ค้นหาใบสำคัญจากเลขที่เอกสาร
        receipts = Receipt.objects.select_related('department', 'created_by').prefetch_related('items').filter(**filters)

        if receipts.count() == 0:
            context['found'] = False
            if dept_code:
                context['error_message'] = f'ไม่พบใบสำคัญรับเงินหมายเลข {receipt_number} ของหน่วยงาน {dept_code}'
            else:
                context['error_message'] = f'ไม่พบใบสำคัญรับเงินหมายเลข {receipt_number}'
        elif receipts.count() == 1:
            # มีใบเดียว แสดงตามปกติ
            context['receipt'] = receipts.first()
            context['found'] = True
            context['qr_url'] = request.build_absolute_uri()
        else:
            # มีหลายใบ (หลายหน่วยงาน) ให้เลือก
            context['found'] = True
            context['multiple_receipts'] = list(receipts)
            context['receipt'] = None
            context['qr_url'] = request.build_absolute_uri()

    except Exception as e:
        context['found'] = False
        context['error_message'] = f'เกิดข้อผิดพลาด: {str(e)}'

    return render(request, 'accounts/receipt_check_public.html', context)


# =============================================================================
# EDIT REQUEST VIEWS
# =============================================================================

@login_required
def edit_request_create_view(request, receipt_id):
    """
    ฟอร์มสำหรับส่งคำร้องขอแก้ไขใบสำคัญรับเงิน
    สำหรับ Basic User ที่ต้องการแก้ไขใบสำคัญ
    """
    receipt = get_object_or_404(Receipt, id=receipt_id)
    
    # ตรวจสอบสิทธิ์ - เฉพาะเจ้าของเอกสารเท่านั้นที่ขอแก้ไขได้
    if receipt.created_by != request.user:
        messages.error(request, 'คุณสามารถขอแก้ไขได้เฉพาะเอกสารที่คุณสร้างเองเท่านั้น')
        return redirect('receipt_detail', receipt_id=receipt.id)
    
    # ตรวจสอบสถานะใบสำคัญ - ต้องเป็น completed เท่านั้น
    if receipt.status != 'completed':
        messages.error(request, 'สามารถขอแก้ไขได้เฉพาะใบสำคัญที่เสร็จสมบูรณ์แล้วเท่านั้น')
        return redirect('receipt_detail', receipt_id=receipt.id)
    
    # ตรวจสอบว่ามีคำร้องขอแก้ไขค้างอยู่หรือไม่
    pending_request = receipt.edit_requests.filter(status='pending').first()
    if pending_request:
        messages.warning(request, f'มีคำร้องขอแก้ไข {pending_request.request_number} รออนุมัติอยู่แล้ว')
        return redirect('edit_request_detail', request_id=pending_request.id)
    
    # Import formset
    from .forms import ReceiptEditRequestItemFormSet
    
    if request.method == 'POST':
        form = ReceiptEditRequestForm(request.POST, receipt=receipt)
        formset = ReceiptEditRequestItemFormSet(request.POST, queryset=receipt.items.all())
        
        # Debug output
        print("Form valid:", form.is_valid())
        if not form.is_valid():
            print("Form errors:", form.errors)
        print("Formset valid:", formset.is_valid())
        if not formset.is_valid():
            print("Formset errors:", formset.errors)
        
        if form.is_valid() and formset.is_valid():
            edit_request = form.save(commit=False)
            edit_request.receipt = receipt
            edit_request.requested_by = request.user
            
            # Calculate new total from formset
            new_total = 0
            for form_item in formset:
                if form_item.cleaned_data and not form_item.cleaned_data.get('DELETE', False):
                    quantity = form_item.cleaned_data.get('quantity', 0)
                    unit_price = form_item.cleaned_data.get('unit_price', 0)
                    new_total += quantity * unit_price
            
            edit_request.new_total_amount = new_total
            edit_request.save()
            
            # Save formset data as JSON for the edit request
            items_data = []
            for form_item in formset:
                if form_item.cleaned_data and not form_item.cleaned_data.get('DELETE', False):
                    items_data.append({
                        'id': form_item.cleaned_data.get('id').id if form_item.cleaned_data.get('id') else None,
                        'description': form_item.cleaned_data.get('description'),
                        'quantity': form_item.cleaned_data.get('quantity'),
                        'unit_price': float(form_item.cleaned_data.get('unit_price', 0)),
                    })
            
            # Store items data in a separate field (we'll add this to model later)
            import json
            edit_request.new_items_data = json.dumps(items_data)
            edit_request.save()
            
            # บันทึก Change Log
            ReceiptChangeLog.log_change(
                receipt=receipt,
                action='edit_requested',
                user=request.user,
                notes=f'ส่งคำร้องขอแก้ไข: {edit_request.reason}',
                edit_request=edit_request
            )
            
            messages.success(request, f'ส่งคำร้องขอแก้ไข {edit_request.request_number} เรียบร้อยแล้ว')
            return redirect('edit_request_detail', request_id=edit_request.id)
    else:
        form = ReceiptEditRequestForm(receipt=receipt)
        formset = ReceiptEditRequestItemFormSet(queryset=receipt.items.all())
    
    context = {
        'form': form,
        'formset': formset,
        'receipt': receipt,
        'page_title': f'ขอแก้ไขใบสำคัญ {receipt.receipt_number}'
    }
    return render(request, 'accounts/edit_request_create.html', context)


@login_required
def edit_request_list_view(request):
    """
    รายการคำร้องขอแก้ไขใบสำคัญรับเงิน
    - Basic User: ดูเฉพาะคำร้องของตัวเอง
    - Department Manager: ดูคำร้องของหน่วยงานตัวเอง
    - Admin: ดูทั้งหมด
    """
    # กรองข้อมูลตามสิทธิ์
    if request.user.has_permission('receipt_view_all'):
        # Admin: ดูทั้งหมด
        edit_requests = ReceiptEditRequest.objects.all()
    elif request.user.has_permission('receipt_edit_approve'):
        # Department Manager: ดูเฉพาะหน่วยงานตัวเอง
        edit_requests = ReceiptEditRequest.objects.filter(
            receipt__department__name=request.user.department
        )
    else:
        # Basic User: ดูเฉพาะของตัวเอง
        edit_requests = ReceiptEditRequest.objects.filter(
            requested_by=request.user
        )
    
    # Filter และ Search
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('q', '')
    
    if status_filter:
        edit_requests = edit_requests.filter(status=status_filter)
    
    if search_query:
        edit_requests = edit_requests.filter(
            Q(request_number__icontains=search_query) |
            Q(receipt__receipt_number__icontains=search_query) |
            Q(receipt__recipient_name__icontains=search_query) |
            Q(reason__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(edit_requests.order_by('-created_at'), 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # สถิติสำหรับ cards
    stats = {
        'total': edit_requests.count(),
        'pending': edit_requests.filter(status='pending').count(),
        'approved': edit_requests.filter(status='approved').count(),
        'rejected': edit_requests.filter(status='rejected').count(),
    }
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'status_filter': status_filter,
        'search_query': search_query,
        'page_title': 'รายการคำร้องขอแก้ไข'
    }
    return render(request, 'accounts/edit_request_list.html', context)


@login_required
def edit_request_detail_view(request, request_id):
    """
    รายละเอียดคำร้องขอแก้ไขใบสำคัญรับเงิน
    รองรับทั้งการดู (GET) และการอนุมัติ/ปฏิเสธ (POST)
    """
    edit_request = get_object_or_404(ReceiptEditRequest, id=request_id)

    # ตรวจสอบสิทธิ์การดู
    can_view = False
    if request.user.has_permission('receipt_view_all'):
        can_view = True
    elif request.user.has_permission('receipt_edit_approve') and edit_request.receipt.department.name == request.user.department:
        can_view = True
    elif edit_request.requested_by == request.user:
        can_view = True

    if not can_view:
        messages.error(request, 'คุณไม่มีสิทธิ์ดูคำร้องนี้')
        return redirect('edit_request_list')

    # ตรวจสอบสิทธิ์การอนุมัติ
    can_approve = edit_request.can_be_approved_by(request.user) and edit_request.status == 'pending'

    # Handle POST request (approval/rejection)
    if request.method == 'POST' and can_approve:
        action = request.POST.get('action')
        notes = request.POST.get('notes', '')

        if action == 'approve':
            edit_request.approve(request.user, notes)

            # บันทึก Change Log
            ReceiptChangeLog.log_change(
                receipt=edit_request.receipt,
                action='edit_approved',
                user=request.user,
                notes=notes,
                edit_request=edit_request
            )

            messages.success(request, f'อนุมัติคำร้อง {edit_request.request_number} เรียบร้อยแล้ว')
            return redirect('receipt_detail', receipt_id=edit_request.receipt.id)

        elif action == 'reject':
            if not notes.strip():
                messages.error(request, 'กรุณาระบุเหตุผลในการปฏิเสธ')
            else:
                edit_request.reject(request.user, notes)

                # บันทึก Change Log
                ReceiptChangeLog.log_change(
                    receipt=edit_request.receipt,
                    action='edit_rejected',
                    user=request.user,
                    notes=notes,
                    edit_request=edit_request
                )

                messages.success(request, f'ปฏิเสธคำร้อง {edit_request.request_number} เรียบร้อยแล้ว')
                return redirect('receipt_detail', receipt_id=edit_request.receipt.id)

    # ประวัติการเปลี่ยนแปลง
    change_logs = edit_request.change_logs.all().order_by('-created_at')

    # Parse new items data if exists
    new_items = []
    if edit_request.new_items_data:
        import json
        try:
            items_data = json.loads(edit_request.new_items_data)
            for item_data in items_data:
                quantity = item_data.get('quantity', 1)
                unit_price = item_data.get('unit_price', 0)
                new_items.append({
                    'description': item_data.get('description', ''),
                    'quantity': quantity,
                    'unit_price': unit_price,
                    'total': quantity * unit_price
                })
        except json.JSONDecodeError:
            pass

    context = {
        'edit_request': edit_request,
        'receipt': edit_request.receipt,
        'can_approve': can_approve,
        'change_logs': change_logs,
        'new_items': new_items,
        'page_title': f'คำร้องขอแก้ไข {edit_request.request_number}'
    }
    return render(request, 'accounts/edit_request_detail.html', context)


@login_required
def edit_request_approval_view(request, request_id):
    """
    หน้าอนุมัติ/ปฏิเสธคำร้องขอแก้ไขใบสำคัญรับเงิน
    สำหรับ Department Manager หรือ Admin
    """
    edit_request = get_object_or_404(ReceiptEditRequest, id=request_id)
    
    # ตรวจสอบสิทธิ์การอนุมัติ
    if not edit_request.can_be_approved_by(request.user):
        messages.error(request, 'คุณไม่มีสิทธิ์อนุมัติคำร้องนี้')
        return redirect('edit_request_detail', request_id=edit_request.id)
    
    # ตรวจสอบสถานะ - ต้องเป็น pending เท่านั้น
    if edit_request.status != 'pending':
        messages.error(request, 'คำร้องนี้ได้รับการพิจารณาแล้ว')
        return redirect('edit_request_detail', request_id=edit_request.id)
    
    if request.method == 'POST':
        form = EditRequestApprovalForm(request.POST)
        if form.is_valid():
            action = form.cleaned_data['action']
            notes = form.cleaned_data['notes']
            
            if action == 'approve':
                edit_request.approve(request.user, notes)
                
                # บันทึก Change Log
                ReceiptChangeLog.log_change(
                    receipt=edit_request.receipt,
                    action='edit_approved',
                    user=request.user,
                    notes=notes,
                    edit_request=edit_request
                )
                
                messages.success(request, f'อนุมัติคำร้อง {edit_request.request_number} เรียบร้อยแล้ว')
                
            elif action == 'reject':
                edit_request.reject(request.user, notes)
                
                # บันทึก Change Log
                ReceiptChangeLog.log_change(
                    receipt=edit_request.receipt,
                    action='edit_rejected',
                    user=request.user,
                    notes=notes,
                    edit_request=edit_request
                )
                
                messages.success(request, f'ปฏิเสธคำร้อง {edit_request.request_number} เรียบร้อยแล้ว')
            
            return redirect('edit_request_detail', request_id=edit_request.id)
    else:
        form = EditRequestApprovalForm()
    
    context = {
        'form': form,
        'edit_request': edit_request,
        'receipt': edit_request.receipt,
        'page_title': f'พิจารณาคำร้อง {edit_request.request_number}'
    }
    return render(request, 'accounts/edit_request_approval.html', context)


@login_required
def edit_request_withdraw_view(request, request_id):
    """
    ถอนคำร้องขอแก้ไขใบสำคัญรับเงิน (สำหรับผู้ส่งคำร้อง)
    """
    edit_request = get_object_or_404(ReceiptEditRequest, id=request_id)
    
    # ตรวจสอบสิทธิ์ - เฉพาะผู้ส่งคำร้อง
    if edit_request.requested_by != request.user:
        messages.error(request, 'คุณไม่มีสิทธิ์ถอนคำร้องนี้')
        return redirect('edit_request_detail', request_id=edit_request.id)
    
    # ตรวจสอบสถานะ - ถอนได้เฉพาะ pending
    if edit_request.status != 'pending':
        messages.error(request, 'สามารถถอนได้เฉพาะคำร้องที่รออนุมัติเท่านั้น')
        return redirect('edit_request_detail', request_id=edit_request.id)
    
    # ถ้าเป็น GET request ให้ redirect ไปยัง detail page โดยตรง
    # เพราะ button มี onclick confirm แล้ว
    if request.method != 'POST':
        return redirect('edit_request_detail', request_id=edit_request.id)
        
    # Process POST request for withdrawal
    edit_request.withdraw()
    
    # บันทึก Change Log
    ReceiptChangeLog.log_change(
        receipt=edit_request.receipt,
        action='edit_requested',  # ใช้ action เดิมเนื่องจากไม่มี 'edit_withdrawn'
        user=request.user,
        notes='ถอนคำร้องขอแก้ไข',
        edit_request=edit_request
    )
    
    messages.success(request, f'ถอนคำร้อง {edit_request.request_number} เรียบร้อยแล้ว')
    return redirect('edit_request_detail', request_id=edit_request.id)


# ===== RECEIPT CANCELLATION SECTION =====

@login_required
def receipt_cancel_direct_view(request, receipt_id):
    """ยกเลิกใบสำคัญโดยตรง (สำหรับ draft หรือ user ที่มีสิทธิ์)"""
    from django.core.exceptions import PermissionDenied
    from django.contrib import messages
    from accounts.models import Receipt
    
    receipt = get_object_or_404(Receipt, id=receipt_id)
    
    # ตรวจสอบว่าสามารถยกเลิกได้โดยตรงหรือไม่
    if not receipt.can_be_cancelled_directly(request.user):
        messages.error(request, 'คุณไม่สามารถยกเลิกใบสำคัญนี้ได้โดยตรง กรุณาส่งคำขออนุมัติ')
        return redirect('receipt_detail', receipt_id=receipt.id)
    
    if request.method == 'POST':
        cancel_reason = request.POST.get('cancel_reason', '')
        
        try:
            receipt.cancel(request.user, cancel_reason)
            messages.success(request, f'ยกเลิกใบสำคัญ {receipt.receipt_number} เรียบร้อยแล้ว')
            return redirect('receipt_detail', receipt_id=receipt.id)
            
        except (PermissionDenied, ValueError) as e:
            messages.error(request, str(e))
            return redirect('receipt_detail', receipt_id=receipt.id)
    
    # GET request - redirect กลับไปหน้า detail
    return redirect('receipt_detail', receipt_id=receipt.id)


@login_required  
def receipt_cancel_request_view(request, receipt_id):
    """สร้างคำขอยกเลิกใบสำคัญ (สำหรับ completed receipts)"""
    from django.contrib import messages
    from accounts.models import Receipt, ReceiptCancelRequest
    
    receipt = get_object_or_404(Receipt, id=receipt_id)
    
    # ตรวจสอบว่าสามารถส่งคำขอยกเลิกได้หรือไม่
    if not receipt.can_be_cancelled_by(request.user):
        messages.error(request, 'คุณไม่มีสิทธิ์ยกเลิกใบสำคัญนี้')
        return redirect('receipt_detail', receipt_id=receipt.id)
    
    # ตรวจสอบสถานะใบสำคัญ
    if receipt.status != 'completed':
        messages.error(request, 'สามารถส่งคำขอยกเลิกได้เฉพาะใบสำคัญที่เสร็จสิ้นแล้วเท่านั้น')
        return redirect('receipt_detail', receipt_id=receipt.id)
        
    # ตรวจสอบว่ามีคำขอรออยู่หรือไม่
    existing_request = ReceiptCancelRequest.objects.filter(
        receipt=receipt,
        status='pending'
    ).first()
    
    if existing_request:
        messages.warning(request, 'มีคำขอยกเลิกรออนุมัติอยู่แล้ว')
        return redirect('cancel_request_detail', request_id=existing_request.id)
    
    if request.method == 'POST':
        cancel_reason = request.POST.get('cancel_reason', '').strip()
        
        if not cancel_reason:
            messages.error(request, 'กรุณาระบุเหตุผลการยกเลิก')
            return render(request, 'accounts/receipt_cancel_request.html', {
                'receipt': receipt,
                'title': f'ขออนุมัติยกเลิก {receipt.receipt_number}'
            })
        
        # สร้างคำขอยกเลิก
        cancel_request = ReceiptCancelRequest.objects.create(
            receipt=receipt,
            requested_by=request.user,
            cancel_reason=cancel_reason
        )
        
        messages.success(request, f'ส่งคำขอยกเลิก {cancel_request.request_number} เรียบร้อยแล้ว')
        return redirect('cancel_request_detail', request_id=cancel_request.id)
    
    # GET request - แสดงฟอร์มขอยกเลิก
    context = {
        'receipt': receipt,
        'title': f'ขออนุมัติยกเลิก {receipt.receipt_number}'
    }
    return render(request, 'accounts/receipt_cancel_request.html', context)


@login_required
def cancel_request_detail_view(request, request_id):
    """แสดงรายละเอียดคำขอยกเลิก และรองรับการอนุมัติ/ปฏิเสธ"""
    from django.core.exceptions import PermissionDenied
    from accounts.models import ReceiptCancelRequest

    cancel_request = get_object_or_404(ReceiptCancelRequest, id=request_id)

    # ตรวจสอบสิทธิ์ดู (เจ้าของคำขอ หรือ ผู้ที่สามารถอนุมัติได้)
    can_view = (
        cancel_request.requested_by == request.user or
        cancel_request.can_be_approved_by(request.user) or
        request.user.has_permission('receipt_view_all')
    )

    if not can_view:
        messages.error(request, 'คุณไม่มีสิทธิ์ดูคำขอยกเลิกนี้')
        return redirect('receipt_list')

    # ตรวจสอบว่าสามารถอนุมัติได้หรือไม่
    can_approve = cancel_request.can_be_approved_by(request.user)

    # Handle POST request สำหรับการอนุมัติ/ปฏิเสธ
    if request.method == 'POST' and can_approve:
        action = request.POST.get('action')
        notes = request.POST.get('notes', '')

        try:
            if action == 'approve':
                cancel_request.approve(request.user, notes)
                messages.success(request, f'อนุมัติคำขอยกเลิก {cancel_request.request_number} เรียบร้อยแล้ว')
            elif action == 'reject':
                cancel_request.reject(request.user, notes)
                messages.success(request, f'ปฏิเสธคำขอยกเลิก {cancel_request.request_number} เรียบร้อยแล้ว')
            else:
                messages.error(request, 'กรุณาเลือกการดำเนินการ')

        except (PermissionDenied, ValueError) as e:
            messages.error(request, str(e))

        return redirect('cancel_request_detail', request_id=cancel_request.id)

    # เพิ่ม can_withdraw ให้ context
    can_withdraw = (
        cancel_request.status == 'pending' and
        cancel_request.requested_by == request.user
    )

    context = {
        'cancel_request': cancel_request,
        'can_approve': can_approve,
        'can_withdraw': can_withdraw,
        'title': f'คำขอยกเลิก {cancel_request.request_number}'
    }
    return render(request, 'accounts/cancel_request_detail.html', context)


# DEPRECATED: ฟอร์มอนุมัติถูกรวมเข้าไปใน cancel_request_detail_view แล้ว
# @login_required
# def cancel_request_approve_view(request, request_id):
#     """อนุมัติคำขอยกเลิก"""
#     from django.contrib import messages
#     from django.core.exceptions import PermissionDenied
#     from accounts.models import ReceiptCancelRequest
#
#     cancel_request = get_object_or_404(ReceiptCancelRequest, id=request_id)
#
#     # ตรวจสอบสิทธิ์
#     if not cancel_request.can_be_approved_by(request.user):
#         messages.error(request, 'คุณไม่มีสิทธิ์อนุมัติคำขอนี้')
#         return redirect('cancel_request_detail', request_id=cancel_request.id)
#
#     if request.method == 'POST':
#         action = request.POST.get('action')
#         approval_notes = request.POST.get('approval_notes', '')
#
#         try:
#             if action == 'approve':
#                 cancel_request.approve(request.user, approval_notes)
#                 messages.success(request, f'อนุมัติคำขอยกเลิก {cancel_request.request_number} เรียบร้อยแล้ว')
#             elif action == 'reject':
#                 cancel_request.reject(request.user, approval_notes)
#                 messages.success(request, f'ปฏิเสธคำขอยกเลิก {cancel_request.request_number} เรียบร้อยแล้ว')
#
#         except (PermissionDenied, ValueError) as e:
#             messages.error(request, str(e))
#
#         return redirect('cancel_request_detail', request_id=cancel_request.id)
#
#     # GET request - แสดงหน้าอนุมัติ
#     context = {
#         'cancel_request': cancel_request,
#         'title': f'พิจารณาคำขอยกเลิก {cancel_request.request_number}'
#     }
#     return render(request, 'accounts/cancel_request_approval.html', context)


@login_required
def cancel_request_withdraw_view(request, request_id):
    """ถอนคำขอยกเลิก"""
    from django.contrib import messages
    from accounts.models import ReceiptCancelRequest
    
    cancel_request = get_object_or_404(ReceiptCancelRequest, id=request_id)
    
    # ตรวจสอบสิทธิ์ - เฉพาะผู้ส่งคำขอ
    if cancel_request.requested_by != request.user:
        messages.error(request, 'คุณไม่มีสิทธิ์ถอนคำขอนี้')
        return redirect('cancel_request_detail', request_id=cancel_request.id)
    
    # ตรวจสอบสถานะ - ถอนได้เฉพาะ pending
    if cancel_request.status != 'pending':
        messages.error(request, 'สามารถถอนได้เฉพาะคำขอที่รออนุมัติเท่านั้น')
        return redirect('cancel_request_detail', request_id=cancel_request.id)
    
    if request.method == 'POST':
        cancel_request.withdraw()
        messages.success(request, f'ถอนคำขอ {cancel_request.request_number} เรียบร้อยแล้ว')
        return redirect('cancel_request_detail', request_id=cancel_request.id)
    
    # GET request - redirect กลับหน้า detail
    return redirect('cancel_request_detail', request_id=cancel_request.id)


# ===== REPORTS SECTION =====

@login_required
def reports_dashboard_view(request):
    """
    รายงานแดชบอร์ด - หน้าหลักของระบบรายงาน
    Features:
    - ยอดรวมทั้งปีงบประมาณ
    - ยอดเดือนนี้
    - ใบสำคัญเสร็จสิ้น
    - คำขอแก้ไข (รอ/อนุมัติ)
    """
    from django.db.models import Sum, Count
    from django.utils import timezone
    from datetime import datetime, timedelta
    from utils.fiscal_year import get_current_fiscal_year, get_fiscal_year_dates
    
    # กำหนด scope การดู (ไม่ต้องเช็ค report_view permission)
    if request.user.has_permission('receipt_view_all'):
        receipts = Receipt.objects.all()
        edit_requests = ReceiptEditRequest.objects.all()
        view_scope = "ทุกหน่วยงาน"
    else:
        # Basic User และ Department Manager - ดูระดับหน่วยงาน
        receipts = Receipt.objects.filter(department__name=request.user.department)
        edit_requests = ReceiptEditRequest.objects.filter(receipt__department__name=request.user.department)
        view_scope = f"หน่วยงาน: {request.user.department}"
    
    # วันที่ปัจจุบัน
    now = timezone.now()
    current_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # ปีงบประมาณปัจจุบัน
    current_fiscal_year = get_current_fiscal_year()
    fiscal_start, fiscal_end = get_fiscal_year_dates(current_fiscal_year)
    fiscal_receipts = receipts.filter(receipt_date__gte=fiscal_start, receipt_date__lte=fiscal_end)
    
    # สถิติเดือนปัจจุบัน
    monthly_receipts = receipts.filter(created_at__gte=current_month_start)
    
    # 1. ยอดรวมทั้งปีงบประมาณ
    fiscal_year_amount = fiscal_receipts.filter(status='completed').aggregate(
        total=Sum('total_amount')
    )['total'] or 0
    fiscal_year_count = fiscal_receipts.filter(status='completed').count()
    
    # 2. ยอดเดือนนี้
    monthly_amount = monthly_receipts.filter(status='completed').aggregate(
        total=Sum('total_amount')
    )['total'] or 0
    monthly_count = monthly_receipts.filter(status='completed').count()
    
    # 3. ใบสำคัญเสร็จสิ้น (ทั้งหมด)
    completed_amount = receipts.filter(status='completed').aggregate(
        total=Sum('total_amount')
    )['total'] or 0
    completed_count = receipts.filter(status='completed').count()
    
    # 4. คำขอแก้ไข
    pending_edit_requests = edit_requests.filter(status='pending').count()
    approved_edit_requests = edit_requests.filter(status__in=['approved', 'applied']).count()
    
    # สถิติสถานะรวม (ใบสำคัญ + คำขอแก้ไข)
    status_summary = []
    
    # สถานะใบสำคัญรับเงิน
    for status_code, status_name in Receipt.STATUS_CHOICES:
        count = receipts.filter(status=status_code).count()
        amount = receipts.filter(status=status_code).aggregate(
            total=Sum('total_amount')
        )['total'] or 0
        
        # กำหนดสี badge
        if status_code == 'completed':
            badge_color = 'success'
        elif status_code == 'draft':
            badge_color = 'warning'
        else:  # cancelled
            badge_color = 'danger'
            
        status_summary.append({
            'name': status_name,
            'count': count,
            'amount': amount,
            'badge_color': badge_color,
            'unit': 'ใบ',
            'type': 'receipt'
        })
    
    # สถานะคำขอแก้ไข
    edit_request_statuses = [
        ('pending', 'รออนุมัติ', 'warning'),
        ('approved', 'อนุมัติแล้ว', 'success'), 
        ('applied', 'ดำเนินการแล้ว', 'info'),
        ('rejected', 'ปฏิเสธ', 'danger'),
        ('withdrawn', 'ถอนคำร้อง', 'secondary'),
    ]
    
    for status_code, status_name, badge_color in edit_request_statuses:
        count = edit_requests.filter(status=status_code).count()
        status_summary.append({
            'name': status_name,
            'count': count,
            'amount': None,  # คำขอแก้ไขไม่มียอดเงิน
            'badge_color': badge_color,
            'unit': 'คำขอ',
            'type': 'edit_request'
        })
    
    # สถิติตามสถานะ (สำหรับส่วนแสดงผล)
    status_stats = {}
    for status_code, status_name in Receipt.STATUS_CHOICES:
        count = receipts.filter(status=status_code).count()
        amount = receipts.filter(status=status_code).aggregate(
            total=Sum('total_amount')
        )['total'] or 0
        status_stats[status_code] = {
            'name': status_name,
            'count': count,
            'amount': amount
        }
    
    # สถิติตามหน่วยงาน (ถ้าดูได้ทุกหน่วยงาน)
    department_stats = []
    if request.user.has_permission('receipt_view_all'):
        departments = Department.objects.filter(is_active=True)
        for dept in departments:
            dept_receipts = receipts.filter(department=dept, status='completed')
            count = dept_receipts.count()
            amount = dept_receipts.aggregate(total=Sum('total_amount'))['total'] or 0
            if count > 0:  # แสดงเฉพาะหน่วยงานที่มีใบสำคัญ
                department_stats.append({
                    'department': dept.name,
                    'count': count,
                    'amount': amount
                })
        
        # เรียงตามยอดเงิน
        department_stats.sort(key=lambda x: x['amount'], reverse=True)
    
    # สถิติ 7 วันที่ผ่านมา (วันปัจจุบันบนสุด)
    daily_stats = []
    day_names_short = ['M', 'Tu', 'W', 'Th', 'F', 'Sa', 'Su']  # 0=Monday, 6=Sunday
    # สีโทนอ่อนสำหรับแต่ละวัน
    day_colors = ['primary', 'success', 'info', 'warning', 'danger', 'secondary', 'dark']
    
    for i in range(6, -1, -1):  # เรียงจาก วันปัจจุบัน ย้อนกลับ 7 วัน
        day = now - timedelta(days=i)
        day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        
        day_receipts = receipts.filter(
            created_at__gte=day_start,
            created_at__lt=day_end,
            status='completed'
        )
        
        # ตัวย่อวันและสี badge
        day_name_short = day_names_short[day.weekday()]
        day_color = day_colors[day.weekday()]
        
        daily_stats.append({
            'day_short': day_name_short,
            'day_color': day_color,
            'date': day.strftime('%d/%m/%y'),  # กลับเป็นตัวเลขเหมือนเดิม
            'count': day_receipts.count(),
            'amount': day_receipts.aggregate(total=Sum('total_amount'))['total'] or 0
        })
    
    # เดือนไทยเต็มสำหรับหัวตาราง
    thai_months_full = [
        'มกราคม', 'กุมภาพันธ์', 'มีนาคม', 'เมษายน', 'พฤษภาคม', 'มิถุนายน',
        'กรกฎาคม', 'สิงหาคม', 'กันยายน', 'ตุลาคม', 'พฤศจิกายน', 'ธันวาคม'
    ]
    current_thai_month = thai_months_full[now.month - 1]
    
    context = {
        'title': 'รายงานแดชบอร์ด',
        'view_scope': view_scope,
        # 4 การ์ดหลัก
        'fiscal_year_amount': fiscal_year_amount,
        'fiscal_year_count': fiscal_year_count,
        'monthly_amount': monthly_amount,
        'monthly_count': monthly_count,
        'completed_amount': completed_amount,
        'completed_count': completed_count,
        'pending_edit_requests': pending_edit_requests,
        'approved_edit_requests': approved_edit_requests,
        # ข้อมูลเสริม
        'current_fiscal_year': current_fiscal_year,
        'status_stats': status_stats,
        'department_stats': department_stats,
        'daily_stats': daily_stats,
        'current_thai_month': current_thai_month,
        'status_summary': status_summary,
        'current_month': now.strftime('%B %Y'),
        'current_month_thai': now.strftime('%m/%Y'),
    }
    
    return render(request, 'accounts/reports_dashboard.html', context)


@login_required
def receipt_report_view(request):
    """
    รายงานใบสำคัญรับเงินแบบรายละเอียด
    Features:
    - Filter ตามวันที่
    - Filter ตามหน่วยงาน
    - Filter ตามสถานะ
    - แสดงเป็นตาราง
    - Export PDF/Excel
    """
    from django.db.models import Sum, Q
    from datetime import datetime
    
    # กำหนด scope การดู (ไม่ต้องเช็ค report_view permission)
    if request.user.has_permission('receipt_view_all'):
        receipts = Receipt.objects.all()
        departments = Department.objects.filter(is_active=True)
        view_scope = "ทุกหน่วยงาน"
    else:
        # Basic User และ Department Manager - ดูระดับหน่วยงาน
        receipts = Receipt.objects.filter(department__name=request.user.department)
        departments = Department.objects.filter(name=request.user.department, is_active=True)
        view_scope = f"หน่วยงาน: {request.user.department}"
    
    # รับค่า filter จาก form
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    department_filter = request.GET.get('department')
    status_filter = request.GET.get('status')
    search_query = request.GET.get('q', '').strip()
    
    # กรอง filter
    filter_applied = False
    
    # Filter วันที่
    if date_from:
        try:
            receipts = receipts.filter(receipt_date__gte=datetime.strptime(date_from, '%Y-%m-%d').date())
            filter_applied = True
        except ValueError:
            messages.warning(request, 'รูปแบบวันที่ไม่ถูกต้อง')
    
    if date_to:
        try:
            receipts = receipts.filter(receipt_date__lte=datetime.strptime(date_to, '%Y-%m-%d').date())
            filter_applied = True
        except ValueError:
            messages.warning(request, 'รูปแบบวันที่ไม่ถูกต้อง')
    
    # Filter หน่วยงาน (ถ้ามีสิทธิ์ดูทุกหน่วยงาน)
    if department_filter and request.user.has_permission('receipt_view_all'):
        receipts = receipts.filter(department__name=department_filter)
        filter_applied = True
    
    # Filter สถานะ
    if status_filter:
        receipts = receipts.filter(status=status_filter)
        filter_applied = True
    
    # ค้นหา
    if search_query:
        receipts = receipts.filter(
            Q(receipt_number__icontains=search_query) |
            Q(recipient_name__icontains=search_query)
        )
        filter_applied = True
    
    # เรียงลำดับ
    receipts = receipts.select_related('department', 'created_by').order_by('-created_at')
    
    # สรุปยอดรวม
    total_amount = receipts.filter(status='completed').aggregate(
        total=Sum('total_amount')
    )['total'] or 0
    
    total_count = receipts.count()
    completed_count = receipts.filter(status='completed').count()
    
    # Pagination
    paginator = Paginator(receipts, 50)  # 50 รายการต่อหน้า
    page = request.GET.get('page')
    receipts_page = paginator.get_page(page)
    
    # สรุปตามสถานะ
    status_summary = {}
    for status_code, status_name in Receipt.STATUS_CHOICES:
        count = receipts.filter(status=status_code).count()
        amount = receipts.filter(status=status_code).aggregate(
            total=Sum('total_amount')
        )['total'] or 0
        status_summary[status_code] = {
            'name': status_name,
            'count': count,
            'amount': amount
        }
    
    context = {
        'title': 'รายงานใบสำคัญรับเงิน',
        'receipts': receipts_page,
        'departments': departments,
        'status_choices': Receipt.STATUS_CHOICES,
        'view_scope': view_scope,
        'total_amount': total_amount,
        'total_count': total_count,
        'completed_count': completed_count,
        'status_summary': status_summary,
        'filter_applied': filter_applied,
        
        # Filter values for form
        'date_from': date_from,
        'date_to': date_to,
        'department_filter': department_filter,
        'status_filter': status_filter,
        'search_query': search_query,
    }
    
    return render(request, 'accounts/receipt_report.html', context)


@login_required
def revenue_summary_report_view(request):
    """
    รายงานสรุปรายรับ - สรุปยอดรับเงินตามช่วงเวลาและหน่วยงาน
    Features:
    - สรุปตามรายวัน, รายเดือน, รายปีงบประมาณ
    - สรุปตามหน่วยงาน
    - Charts และ graphs
    - Export Excel/PDF
    """
    from django.db.models import Sum, Count
    from django.utils import timezone
    from datetime import datetime, timedelta
    from utils.fiscal_year import get_current_fiscal_year, get_fiscal_year_dates
    
    # กำหนด scope การดู
    if request.user.has_permission('receipt_view_all'):
        receipts = Receipt.objects.all()
        departments = Department.objects.filter(is_active=True)
        view_scope = "ทุกหน่วยงาน"
    else:
        receipts = Receipt.objects.filter(department__name=request.user.department)
        departments = Department.objects.filter(name=request.user.department, is_active=True)
        view_scope = f"หน่วยงาน: {request.user.department}"
    
    # รับค่า filter
    period_type = request.GET.get('period', 'monthly')  # daily, monthly, fiscal_year
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    department_filter = request.GET.get('department')
    
    # ตรวจสอบ Custom Date Range Mode
    is_custom_mode = bool(date_from or date_to)
    
    # Filter หน่วยงาน
    if department_filter and request.user.has_permission('receipt_view_all'):
        receipts = receipts.filter(department__name=department_filter)
    
    # Filter วันที่ (เฉพาะ custom mode)
    if is_custom_mode:
        if date_from:
            try:
                receipts = receipts.filter(receipt_date__gte=datetime.strptime(date_from, '%Y-%m-%d').date())
            except ValueError:
                pass
        
        if date_to:
            try:
                receipts = receipts.filter(receipt_date__lte=datetime.strptime(date_to, '%Y-%m-%d').date())
            except ValueError:
                pass
    
    # เฉพาะใบสำคัญที่เสร็จสิ้น
    completed_receipts = receipts.filter(status='completed')
    
    # สรุปรวมทั้งหมด
    total_summary = {
        'total_amount': completed_receipts.aggregate(total=Sum('total_amount'))['total'] or 0,
        'total_count': completed_receipts.count(),
        'total_departments': completed_receipts.values('department').distinct().count(),
    }
    
    # สรุปตามหน่วยงาน
    department_summary = []
    for dept in departments:
        dept_receipts = completed_receipts.filter(department=dept)
        count = dept_receipts.count()
        amount = dept_receipts.aggregate(total=Sum('total_amount'))['total'] or 0
        
        if count > 0:  # แสดงเฉพาะหน่วยงานที่มีข้อมูล
            department_summary.append({
                'department': dept.name,
                'department_code': dept.code,
                'count': count,
                'amount': amount,
                'percentage': round((amount / total_summary['total_amount'] * 100) if total_summary['total_amount'] > 0 else 0, 1)
            })
    
    # เรียงตามยอดเงิน
    department_summary.sort(key=lambda x: x['amount'], reverse=True)
    
    # สรุปตามช่วงเวลา
    period_summary = []
    now = timezone.now()
    
    if is_custom_mode:
        # โหมดกำหนดเอง - สร้าง period summary แบบรายวันตามช่วงที่กำหนด
        from datetime import datetime, timedelta
        
        # กำหนดช่วงวันที่
        if date_from and date_to:
            start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
        elif date_from:
            start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            end_date = now.date()  # จนถึงวันนี้
        elif date_to:
            end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
            start_date = end_date - timedelta(days=30)  # ย้อนหลัง 30 วัน
        else:
            # ไม่ควรเกิดขึ้น แต่ป้องกันไว้
            start_date = now.date() - timedelta(days=30)
            end_date = now.date()
        
        # สร้างรายการวันที่ในช่วงที่กำหนด
        current_date = start_date
        while current_date <= end_date:
            day_start = datetime.combine(current_date, datetime.min.time())
            day_start = timezone.make_aware(day_start)
            day_end = day_start + timedelta(days=1)
            
            day_receipts = completed_receipts.filter(
                created_at__gte=day_start,
                created_at__lt=day_end
            )
            
            period_summary.append({
                'period': current_date.strftime('%d/%m/%y'),
                'count': day_receipts.count(),
                'amount': day_receipts.aggregate(total=Sum('total_amount'))['total'] or 0
            })
            
            current_date += timedelta(days=1)
    elif period_type == 'daily':
        # รายวัน (30 วันล่าสุด)
        for i in range(29, -1, -1):
            day = now - timedelta(days=i)
            day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
            day_end = day_start + timedelta(days=1)
            
            day_receipts = completed_receipts.filter(
                created_at__gte=day_start,
                created_at__lt=day_end
            )
            
            period_summary.append({
                'period': day.strftime('%d/%m/%y'),
                'count': day_receipts.count(),
                'amount': day_receipts.aggregate(total=Sum('total_amount'))['total'] or 0
            })
    
    elif period_type == 'monthly':
        # รายเดือน (12 เดือนล่าสุด)
        for i in range(11, -1, -1):
            # คำนวณเดือนย้อนหลังอย่างแม่นยำ
            current_month = now.month
            current_year = now.year
            
            target_month = current_month - i
            target_year = current_year
            
            # ปรับปีถ้าเดือนติดลบ
            while target_month <= 0:
                target_month += 12
                target_year -= 1
            
            # วันแรกของเดือนเป้าหมาย
            month_start = now.replace(year=target_year, month=target_month, day=1, hour=0, minute=0, second=0, microsecond=0)
            
            # วันแรกของเดือนถัดไป
            if target_month == 12:
                next_month_start = now.replace(year=target_year + 1, month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            else:
                next_month_start = now.replace(year=target_year, month=target_month + 1, day=1, hour=0, minute=0, second=0, microsecond=0)
            
            month_receipts = completed_receipts.filter(
                created_at__gte=month_start,
                created_at__lt=next_month_start
            )
            
            # เดือนไทย
            thai_months = ['ม.ค.', 'ก.พ.', 'มี.ค.', 'เม.ย.', 'พ.ค.', 'มิ.ย.',
                          'ก.ค.', 'ส.ค.', 'ก.ย.', 'ต.ค.', 'พ.ย.', 'ธ.ค.']
            thai_month = thai_months[target_month - 1]
            
            period_summary.append({
                'period': f"{thai_month} {target_year + 543}",
                'count': month_receipts.count(),
                'amount': month_receipts.aggregate(total=Sum('total_amount'))['total'] or 0
            })
    
    elif period_type == 'fiscal_year':
        # รายปีงบประมาณ (5 ปีล่าสุด)
        current_fiscal = get_current_fiscal_year()
        for i in range(4, -1, -1):
            fiscal_year = current_fiscal - i
            fiscal_start, fiscal_end = get_fiscal_year_dates(fiscal_year)
            
            fiscal_receipts = completed_receipts.filter(
                receipt_date__gte=fiscal_start,
                receipt_date__lte=fiscal_end
            )
            
            period_summary.append({
                'period': f"ปีงบ {fiscal_year}",
                'count': fiscal_receipts.count(),
                'amount': fiscal_receipts.aggregate(total=Sum('total_amount'))['total'] or 0
            })
    
    context = {
        'title': 'รายงานสรุปรายรับ',
        'view_scope': view_scope,
        'total_summary': total_summary,
        'department_summary': department_summary,
        'period_summary': period_summary,
        'period_type': period_type,
        'departments': departments,
        'is_custom_mode': is_custom_mode,
        
        # Filter values
        'date_from': date_from,
        'date_to': date_to,
        'department_filter': department_filter,
    }
    
    return render(request, 'accounts/revenue_summary_report.html', context)


@login_required
def revenue_summary_excel_export(request):
    """
    Export รายงานสรุปรายรับเป็น Excel
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.chart import BarChart, Reference
    from django.http import HttpResponse
    from django.db.models import Sum, Count
    from django.utils import timezone
    from datetime import datetime, timedelta
    from utils.fiscal_year import get_current_fiscal_year, get_fiscal_year_dates
    
    # ใช้ logic เดียวกันกับ revenue_summary_report_view
    if request.user.has_permission('receipt_view_all'):
        receipts = Receipt.objects.all()
        departments = Department.objects.filter(is_active=True)
        view_scope = "ทุกหน่วยงาน"
    else:
        receipts = Receipt.objects.filter(department__name=request.user.department)
        departments = Department.objects.filter(name=request.user.department, is_active=True)
        view_scope = f"หน่วยงาน: {request.user.department}"
    
    # รับค่า filter
    period_type = request.GET.get('period', 'monthly')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    department_filter = request.GET.get('department')
    
    # ตรวจสอบ Custom Date Range Mode
    is_custom_mode = bool(date_from or date_to)
    
    # Apply filters
    if department_filter and request.user.has_permission('receipt_view_all'):
        receipts = receipts.filter(department__name=department_filter)
    
    # Filter วันที่ (เฉพาะ custom mode)
    if is_custom_mode:
        if date_from:
            try:
                receipts = receipts.filter(receipt_date__gte=datetime.strptime(date_from, '%Y-%m-%d').date())
            except ValueError:
                pass
        
        if date_to:
            try:
                receipts = receipts.filter(receipt_date__lte=datetime.strptime(date_to, '%Y-%m-%d').date())
            except ValueError:
                pass
    
    completed_receipts = receipts.filter(status='completed')
    
    # สรุปรวม
    total_summary = {
        'total_amount': completed_receipts.aggregate(total=Sum('total_amount'))['total'] or 0,
        'total_count': completed_receipts.count(),
        'total_departments': completed_receipts.values('department').distinct().count(),
    }
    
    # สรุปตามหน่วยงาน
    department_summary = []
    for dept in departments:
        dept_receipts = completed_receipts.filter(department=dept)
        count = dept_receipts.count()
        amount = dept_receipts.aggregate(total=Sum('total_amount'))['total'] or 0
        
        if count > 0:
            department_summary.append({
                'department': dept.name,
                'department_code': dept.code,
                'count': count,
                'amount': amount,
                'percentage': round((amount / total_summary['total_amount'] * 100) if total_summary['total_amount'] > 0 else 0, 1)
            })
    
    department_summary.sort(key=lambda x: x['amount'], reverse=True)
    
    # สรุปตามช่วงเวลา
    period_summary = []
    now = timezone.now()
    
    if is_custom_mode:
        # โหมดกำหนดเอง - สร้าง period summary แบบรายวันตามช่วงที่กำหนด
        if date_from and date_to:
            start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
        elif date_from:
            start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            end_date = now.date()
        elif date_to:
            end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
            start_date = end_date - timedelta(days=30)
        else:
            start_date = now.date() - timedelta(days=30)
            end_date = now.date()
        
        current_date = start_date
        while current_date <= end_date:
            day_start = datetime.combine(current_date, datetime.min.time())
            day_start = timezone.make_aware(day_start)
            day_end = day_start + timedelta(days=1)
            
            day_receipts = completed_receipts.filter(
                created_at__gte=day_start,
                created_at__lt=day_end
            )
            
            period_summary.append({
                'period': current_date.strftime('%d/%m/%y'),
                'count': day_receipts.count(),
                'amount': day_receipts.aggregate(total=Sum('total_amount'))['total'] or 0
            })
            
            current_date += timedelta(days=1)
    elif period_type == 'daily':
        for i in range(29, -1, -1):
            day = now - timedelta(days=i)
            day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
            day_end = day_start + timedelta(days=1)
            
            day_receipts = completed_receipts.filter(
                created_at__gte=day_start,
                created_at__lt=day_end
            )
            
            period_summary.append({
                'period': day.strftime('%d/%m/%y'),
                'count': day_receipts.count(),
                'amount': day_receipts.aggregate(total=Sum('total_amount'))['total'] or 0
            })
    
    elif period_type == 'monthly':
        for i in range(11, -1, -1):
            # คำนวณเดือนย้อนหลังอย่างแม่นยำ
            current_month = now.month
            current_year = now.year
            
            target_month = current_month - i
            target_year = current_year
            
            # ปรับปีถ้าเดือนติดลบ
            while target_month <= 0:
                target_month += 12
                target_year -= 1
            
            # วันแรกของเดือนเป้าหมาย
            month_start = now.replace(year=target_year, month=target_month, day=1, hour=0, minute=0, second=0, microsecond=0)
            
            # วันแรกของเดือนถัดไป
            if target_month == 12:
                next_month_start = now.replace(year=target_year + 1, month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            else:
                next_month_start = now.replace(year=target_year, month=target_month + 1, day=1, hour=0, minute=0, second=0, microsecond=0)
            
            month_receipts = completed_receipts.filter(
                created_at__gte=month_start,
                created_at__lt=next_month_start
            )
            
            thai_months = ['ม.ค.', 'ก.พ.', 'มี.ค.', 'เม.ย.', 'พ.ค.', 'มิ.ย.',
                          'ก.ค.', 'ส.ค.', 'ก.ย.', 'ต.ค.', 'พ.ย.', 'ธ.ค.']
            thai_month = thai_months[target_month - 1]
            
            period_summary.append({
                'period': f"{thai_month} {target_year + 543}",
                'count': month_receipts.count(),
                'amount': month_receipts.aggregate(total=Sum('total_amount'))['total'] or 0
            })
    
    elif period_type == 'fiscal_year':
        current_fiscal = get_current_fiscal_year()
        for i in range(4, -1, -1):
            fiscal_year = current_fiscal - i
            fiscal_start, fiscal_end = get_fiscal_year_dates(fiscal_year)
            
            fiscal_receipts = completed_receipts.filter(
                receipt_date__gte=fiscal_start,
                receipt_date__lte=fiscal_end
            )
            
            period_summary.append({
                'period': f"ปีงบ {fiscal_year}",
                'count': fiscal_receipts.count(),
                'amount': fiscal_receipts.aggregate(total=Sum('total_amount'))['total'] or 0
            })
    
    # สร้าง Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "รายงานสรุปรายรับ"
    
    # กำหนดสี
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    summary_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header information
    current_fiscal = get_current_fiscal_year()
    ws['A1'] = "รายงานสรุปรายรับ"
    ws['A2'] = f"ประจำปีงบประมาณ {current_fiscal}"
    ws['A3'] = f"ขอบเขต: {view_scope}"
    
    # ช่วงวันที่
    date_range = ""
    if is_custom_mode:
        if date_from and date_to:
            date_range = f"โหมดกำหนดเอง: {date_from} ถึง {date_to}"
        elif date_from:
            date_range = f"โหมดกำหนดเอง: ตั้งแต่วันที่ {date_from}"
        elif date_to:
            date_range = f"โหมดกำหนดเอง: จนถึงวันที่ {date_to}"
    else:
        date_range = f"โหมด: {period_type}"
    
    ws['A4'] = date_range
    
    # ปรับ style สำหรับ header
    for row in range(1, 5):
        ws[f'A{row}'].font = Font(bold=True, size=14 if row == 1 else 12)
        ws[f'A{row}'].alignment = Alignment(horizontal='center' if row == 1 else 'left')
    
    ws.merge_cells('A1:E1')
    
    # สรุปยอดรวม
    ws['A6'] = "สรุปยอดรวม"
    ws['A6'].font = Font(bold=True, size=12)
    
    ws['A7'] = "ยอดเงินรวม:"
    ws['B7'] = f"{total_summary['total_amount']:,.2f} บาท"
    ws['A8'] = "จำนวนใบสำคัญ:"
    ws['B8'] = f"{total_summary['total_count']:,} ใบ"
    ws['A9'] = "หน่วยงานที่มีข้อมูล:"
    ws['B9'] = f"{total_summary['total_departments']} หน่วยงาน"
    
    # สรุปตามช่วงเวลา
    start_row = 11
    if is_custom_mode:
        ws[f'A{start_row}'] = "สรุปตามช่วงเวลา (รายวัน - กำหนดเอง)"
    else:
        ws[f'A{start_row}'] = f"สรุปตามช่วงเวลา ({period_type})"
    ws[f'A{start_row}'].font = Font(bold=True, size=12)
    
    # Headers สำหรับตารางช่วงเวลา
    headers = ['ช่วงเวลา', 'จำนวน (ใบ)', 'ยอดเงิน (บาท)']
    start_row += 1
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # ข้อมูลช่วงเวลา
    for i, period in enumerate(period_summary, 1):
        row = start_row + i
        ws.cell(row=row, column=1, value=period['period']).border = border
        ws.cell(row=row, column=2, value=period['count']).border = border
        ws.cell(row=row, column=3, value=period['amount']).border = border
        
        # จัดรูปแบบตัวเลข
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
    
    # สรุปตามหน่วยงาน
    dept_start_row = start_row + len(period_summary) + 3
    ws[f'A{dept_start_row}'] = "สรุปตามหน่วยงาน"
    ws[f'A{dept_start_row}'].font = Font(bold=True, size=12)
    
    # Headers สำหรับตารางหน่วยงาน
    dept_headers = ['หน่วยงาน', 'รหัส', 'จำนวน (ใบ)', 'ยอดเงิน (บาท)', 'เปอร์เซ็นต์']
    dept_start_row += 1
    
    for col, header in enumerate(dept_headers, 1):
        cell = ws.cell(row=dept_start_row, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # ข้อมูลหน่วยงาน
    for i, dept in enumerate(department_summary, 1):
        row = dept_start_row + i
        ws.cell(row=row, column=1, value=dept['department']).border = border
        ws.cell(row=row, column=2, value=dept['department_code']).border = border
        ws.cell(row=row, column=3, value=dept['count']).border = border
        ws.cell(row=row, column=4, value=dept['amount']).border = border
        ws.cell(row=row, column=5, value=f"{dept['percentage']}%").border = border
        
        # จัดรูปแบบ
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')
    
    # ปรับความกว้างคอลัมน์
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    
    # สร้าง response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="revenue_summary_report.xlsx"'
    
    wb.save(response)
    return response


@login_required
def revenue_summary_pdf_export(request):
    """
    Export รายงานสรุปรายรับเป็น PDF
    """
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from django.http import HttpResponse
    from django.db.models import Sum, Count
    from django.utils import timezone
    from datetime import datetime, timedelta
    from utils.fiscal_year import get_current_fiscal_year, get_fiscal_year_dates
    import os
    
    # ลงทะเบียนฟอนต์ไทย
    font_path = os.path.join(os.path.dirname(__file__), '..', 'static', 'fonts', 'THSarabunNew.ttf')
    if os.path.exists(font_path):
        pdfmetrics.registerFont(TTFont('THSarabunNew', font_path))
        thai_font = 'THSarabunNew'
    else:
        thai_font = 'Helvetica'
    
    # ใช้ logic เดียวกันกับ revenue_summary_report_view
    if request.user.has_permission('receipt_view_all'):
        receipts = Receipt.objects.all()
        departments = Department.objects.filter(is_active=True)
        view_scope = "ทุกหน่วยงาน"
    else:
        receipts = Receipt.objects.filter(department__name=request.user.department)
        departments = Department.objects.filter(name=request.user.department, is_active=True)
        view_scope = f"หน่วยงาน: {request.user.department}"
    
    # รับค่า filter
    period_type = request.GET.get('period', 'monthly')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    department_filter = request.GET.get('department')
    
    # ตรวจสอบ Custom Date Range Mode
    is_custom_mode = bool(date_from or date_to)
    
    # Apply filters
    if department_filter and request.user.has_permission('receipt_view_all'):
        receipts = receipts.filter(department__name=department_filter)
    
    # Filter วันที่ (เฉพาะ custom mode)
    if is_custom_mode:
        if date_from:
            try:
                receipts = receipts.filter(receipt_date__gte=datetime.strptime(date_from, '%Y-%m-%d').date())
            except ValueError:
                pass
        
        if date_to:
            try:
                receipts = receipts.filter(receipt_date__lte=datetime.strptime(date_to, '%Y-%m-%d').date())
            except ValueError:
                pass
    
    completed_receipts = receipts.filter(status='completed')
    
    # สรุปข้อมูล (ใช้ logic เดียวกัน)
    total_summary = {
        'total_amount': completed_receipts.aggregate(total=Sum('total_amount'))['total'] or 0,
        'total_count': completed_receipts.count(),
        'total_departments': completed_receipts.values('department').distinct().count(),
    }
    
    # สรุปตามหน่วยงาน
    department_summary = []
    for dept in departments:
        dept_receipts = completed_receipts.filter(department=dept)
        count = dept_receipts.count()
        amount = dept_receipts.aggregate(total=Sum('total_amount'))['total'] or 0
        
        if count > 0:
            department_summary.append({
                'department': dept.name,
                'department_code': dept.code,
                'count': count,
                'amount': amount,
                'percentage': round((amount / total_summary['total_amount'] * 100) if total_summary['total_amount'] > 0 else 0, 1)
            })
    
    department_summary.sort(key=lambda x: x['amount'], reverse=True)
    
    # สรุปตามช่วงเวลา
    period_summary = []
    now = timezone.now()
    
    if is_custom_mode:
        # โหมดกำหนดเอง - สร้าง period summary แบบรายวันตามช่วงที่กำหนด
        if date_from and date_to:
            start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
        elif date_from:
            start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            end_date = now.date()
        elif date_to:
            end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
            start_date = end_date - timedelta(days=30)
        else:
            start_date = now.date() - timedelta(days=30)
            end_date = now.date()
        
        current_date = start_date
        while current_date <= end_date:
            day_start = datetime.combine(current_date, datetime.min.time())
            day_start = timezone.make_aware(day_start)
            day_end = day_start + timedelta(days=1)
            
            day_receipts = completed_receipts.filter(
                created_at__gte=day_start,
                created_at__lt=day_end
            )
            
            period_summary.append({
                'period': current_date.strftime('%d/%m/%y'),
                'count': day_receipts.count(),
                'amount': day_receipts.aggregate(total=Sum('total_amount'))['total'] or 0
            })
            
            current_date += timedelta(days=1)
    elif period_type == 'daily':
        for i in range(29, -1, -1):
            day = now - timedelta(days=i)
            day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
            day_end = day_start + timedelta(days=1)
            
            day_receipts = completed_receipts.filter(
                created_at__gte=day_start,
                created_at__lt=day_end
            )
            
            period_summary.append({
                'period': day.strftime('%d/%m/%y'),
                'count': day_receipts.count(),
                'amount': day_receipts.aggregate(total=Sum('total_amount'))['total'] or 0
            })
    
    elif period_type == 'monthly':
        for i in range(11, -1, -1):
            # คำนวณเดือนย้อนหลังอย่างแม่นยำ
            current_month = now.month
            current_year = now.year
            
            target_month = current_month - i
            target_year = current_year
            
            # ปรับปีถ้าเดือนติดลบ
            while target_month <= 0:
                target_month += 12
                target_year -= 1
            
            # วันแรกของเดือนเป้าหมาย
            month_start = now.replace(year=target_year, month=target_month, day=1, hour=0, minute=0, second=0, microsecond=0)
            
            # วันแรกของเดือนถัดไป
            if target_month == 12:
                next_month_start = now.replace(year=target_year + 1, month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            else:
                next_month_start = now.replace(year=target_year, month=target_month + 1, day=1, hour=0, minute=0, second=0, microsecond=0)
            
            month_receipts = completed_receipts.filter(
                created_at__gte=month_start,
                created_at__lt=next_month_start
            )
            
            thai_months = ['ม.ค.', 'ก.พ.', 'มี.ค.', 'เม.ย.', 'พ.ค.', 'มิ.ย.',
                          'ก.ค.', 'ส.ค.', 'ก.ย.', 'ต.ค.', 'พ.ย.', 'ธ.ค.']
            thai_month = thai_months[target_month - 1]
            
            period_summary.append({
                'period': f"{thai_month} {target_year + 543}",
                'count': month_receipts.count(),
                'amount': month_receipts.aggregate(total=Sum('total_amount'))['total'] or 0
            })
    
    elif period_type == 'fiscal_year':
        current_fiscal = get_current_fiscal_year()
        for i in range(4, -1, -1):
            fiscal_year = current_fiscal - i
            fiscal_start, fiscal_end = get_fiscal_year_dates(fiscal_year)
            
            fiscal_receipts = completed_receipts.filter(
                receipt_date__gte=fiscal_start,
                receipt_date__lte=fiscal_end
            )
            
            period_summary.append({
                'period': f"ปีงบ {fiscal_year}",
                'count': fiscal_receipts.count(),
                'amount': fiscal_receipts.aggregate(total=Sum('total_amount'))['total'] or 0
            })
    
    # สร้าง PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="revenue_summary_report.pdf"'
    
    # สร้าง PDF document (landscape สำหรับตารางกว้าง)
    from reportlab.lib.pagesizes import landscape
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=0.5*inch)
    
    # เตรียม styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Title'],
        fontName=thai_font,
        fontSize=18,
        alignment=1,
        spaceAfter=20
    )
    
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontName=thai_font,
        fontSize=12,
        alignment=1,
        spaceAfter=10
    )
    
    # สร้างเนื้อหา PDF
    story = []
    
    # Header information
    current_fiscal = get_current_fiscal_year()
    story.append(Paragraph('รายงานสรุปรายรับ', title_style))
    story.append(Paragraph(f'ประจำปีงบประมาณ {current_fiscal}', header_style))
    story.append(Paragraph(f'ขอบเขต: {view_scope}', header_style))
    
    # ช่วงวันที่
    date_range = ""
    if date_from and date_to:
        date_range = f"ระหว่างวันที่ {date_from} ถึง {date_to}"
    elif date_from:
        date_range = f"ตั้งแต่วันที่ {date_from}"
    elif date_to:
        date_range = f"จนถึงวันที่ {date_to}"
    else:
        date_range = "ทุกช่วงเวลา"
    
    story.append(Paragraph(date_range, header_style))
    story.append(Spacer(1, 20))
    
    # สรุปยอดรวม
    summary_data = [
        ['สรุปยอดรวม', ''],
        ['ยอดเงินรวม:', f'{total_summary["total_amount"]:,.2f} บาท'],
        ['จำนวนใบสำคัญ:', f'{total_summary["total_count"]:,} ใบ'],
        ['หน่วยงานที่มีข้อมูล:', f'{total_summary["total_departments"]} หน่วยงาน']
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), thai_font),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (0, 0), (1, 0), colors.lightblue),
        ('FONTNAME', (0, 0), (1, 0), thai_font),
        ('FONTSIZE', (0, 0), (1, 0), 14),
        ('SPAN', (0, 0), (1, 0)),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 1), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # สรุปตามช่วงเวลา
    if period_summary:
        period_title = "แนวโน้มรายรับ"
        if period_type == 'daily':
            period_title = "แนวโน้มรายรับ (รายวัน)"
        elif period_type == 'monthly':
            period_title = "แนวโน้มรายรับ (รายเดือน)"
        elif period_type == 'fiscal_year':
            period_title = "แนวโน้มรายรับ (รายปีงบประมาณ)"
        
        story.append(Paragraph(period_title, ParagraphStyle(
            'SubHeader',
            parent=styles['Heading2'],
            fontName=thai_font,
            fontSize=14,
            spaceAfter=10
        )))
        
        period_headers = ['ช่วงเวลา', 'จำนวน (ใบ)', 'ยอดเงิน (บาท)']
        period_data = [period_headers]
        
        # แสดงข้อมูลครบถ้วนตามที่ผู้ใช้เลือก period (สำหรับการตรวจสอบ)
        for period in period_summary:
            period_data.append([
                period['period'],
                str(period['count']),
                f"{period['amount']:,.0f}"
            ])
        
        period_table = Table(period_data, colWidths=[3*inch, 2*inch, 3*inch])
        period_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), thai_font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        story.append(period_table)
        story.append(Spacer(1, 20))
    
    # สรุปตามหน่วยงาน
    if department_summary:
        story.append(Paragraph('สรุปตามหน่วยงาน', ParagraphStyle(
            'SubHeader',
            parent=styles['Heading2'],
            fontName=thai_font,
            fontSize=14,
            spaceAfter=10
        )))
        
        dept_headers = ['หน่วยงาน', 'รหัส', 'จำนวน (ใบ)', 'ยอดเงิน (บาท)', 'เปอร์เซ็นต์']
        dept_data = [dept_headers]
        
        for dept in department_summary[:10]:  # จำกัด 10 อันดับแรก
            dept_data.append([
                dept['department'][:20],
                dept['department_code'],
                str(dept['count']),
                f"{dept['amount']:,.0f}",
                f"{dept['percentage']}%"
            ])
        
        dept_table = Table(dept_data, colWidths=[2*inch, 1*inch, 1*inch, 1.5*inch, 1*inch])
        dept_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), thai_font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        story.append(dept_table)
    
    # สร้าง PDF
    doc.build(story)
    
    return response


@login_required
def receipt_report_pdf_export(request):
    """
    Export รายงานใบสำคัญรับเงินเป็น PDF
    """
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from django.http import HttpResponse
    from django.db.models import Sum, Q
    from datetime import datetime
    from utils.fiscal_year import get_current_fiscal_year
    import os
    
    # ลงทะเบียนฟอนต์ไทย
    font_path = os.path.join(os.path.dirname(__file__), '..', 'static', 'fonts', 'THSarabunNew.ttf')
    if os.path.exists(font_path):
        pdfmetrics.registerFont(TTFont('THSarabunNew', font_path))
        thai_font = 'THSarabunNew'
    else:
        thai_font = 'Helvetica'  # fallback
    
    # ใช้ logic เดียวกันกับ receipt_report_view สำหรับ filter
    if request.user.has_permission('receipt_view_all'):
        receipts = Receipt.objects.all()
        view_scope = "ทุกหน่วยงาน"
    else:
        receipts = Receipt.objects.filter(department__name=request.user.department)
        view_scope = f"หน่วยงาน: {request.user.department}"
    
    # รับค่า filter
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    department_filter = request.GET.get('department')
    status_filter = request.GET.get('status')
    search_query = request.GET.get('q', '').strip()
    
    # Apply filters
    if date_from:
        try:
            receipts = receipts.filter(receipt_date__gte=datetime.strptime(date_from, '%Y-%m-%d').date())
        except ValueError:
            pass
    
    if date_to:
        try:
            receipts = receipts.filter(receipt_date__lte=datetime.strptime(date_to, '%Y-%m-%d').date())
        except ValueError:
            pass
    
    if department_filter and request.user.has_permission('receipt_view_all'):
        receipts = receipts.filter(department__name=department_filter)
    
    if status_filter:
        receipts = receipts.filter(status=status_filter)
    
    if search_query:
        receipts = receipts.filter(
            Q(receipt_number__icontains=search_query) |
            Q(recipient_name__icontains=search_query)
        )
    
    receipts = receipts.select_related('department', 'created_by').prefetch_related('items').order_by('-created_at')
    
    # สร้าง PDF response (inline - เปิดในแท็บใหม่)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="receipt_report.pdf"'
    
    # สร้าง PDF document
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=0.5*inch)
    
    # เตรียม styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Title'],
        fontName=thai_font,
        fontSize=16,
        alignment=1,  # center
        spaceAfter=20
    )
    
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontName=thai_font,
        fontSize=12,
        alignment=1,
        spaceAfter=10
    )
    
    # สร้างเนื้อหา PDF
    story = []
    
    # Header information
    current_fiscal = get_current_fiscal_year()
    story.append(Paragraph('รายงานใบสำคัญรับเงิน', title_style))
    story.append(Paragraph(f'ประจำปีงบประมาณ {current_fiscal}', header_style))
    
    # ขอบเขต
    if department_filter:
        story.append(Paragraph(f'ชื่อหน่วยงาน: {department_filter}', header_style))
    else:
        story.append(Paragraph(f'ขอบเขต: {view_scope}', header_style))
    
    # ช่วงวันที่
    date_range = ""
    if date_from and date_to:
        date_range = f"ระหว่างวันที่ {date_from} ถึง {date_to}"
    elif date_from:
        date_range = f"ตั้งแต่วันที่ {date_from}"
    elif date_to:
        date_range = f"จนถึงวันที่ {date_to}"
    else:
        date_range = "ทุกช่วงเวลา"
    
    story.append(Paragraph(date_range, header_style))
    story.append(Spacer(1, 20))
    
    # สร้างตาราง
    headers = [
        'ลำดับ', 'ใบสำคัญเลขที่', 'วันที่ขอ', 'รายการ',
        'จำนวนเงิน', 'ผู้รับเงิน', 'ผู้จ่ายเงิน', 'หมายเหตุ'
    ]
    
    table_data = [headers]
    
    for index, receipt in enumerate(receipts[:100], 1):  # จำกัด 100 รายการเพื่อป้องกัน PDF ใหญ่เกินไป
        # รวมรายการ
        items_text = []
        for item in receipt.items.all():
            items_text.append(f"{item.description}")
        items_display = "; ".join(items_text) if items_text else "-"
        
        # ข้อมูลแถว
        payer = receipt.created_by.get_display_name() if receipt.created_by else "-"
        status_map = {'draft': 'ร่าง', 'completed': 'เสร็จสิ้น', 'cancelled': 'ยกเลิก'}
        
        row_data = [
            str(index),
            receipt.receipt_number or "-",
            receipt.receipt_date.strftime('%d/%m/%Y') if receipt.receipt_date else "-",
            items_display[:30] + "..." if len(items_display) > 30 else items_display,
            f"{receipt.total_amount:,.2f}",
            receipt.recipient_name[:20] + "..." if len(receipt.recipient_name) > 20 else receipt.recipient_name,
            payer[:15] + "..." if len(payer) > 15 else payer,
            status_map.get(receipt.status, receipt.status)
        ]
        table_data.append(row_data)
    
    # สร้างตาราง
    table = Table(table_data, colWidths=[0.8*inch, 1.2*inch, 1*inch, 2*inch, 1*inch, 1.5*inch, 1*inch, 0.8*inch])
    table.setStyle(TableStyle([
        # Header style
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), thai_font),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Data style
        ('FONTNAME', (0, 1), (-1, -1), thai_font),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Alternate row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    story.append(table)
    
    # สร้าง PDF
    doc.build(story)
    
    return response


@login_required
def receipt_report_excel_export(request):
    """
    Export รายงานใบสำคัญรับเงินเป็น Excel ตามฟอร์มที่กำหนด
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from django.db.models import Sum, Q
    from datetime import datetime
    from utils.fiscal_year import get_current_fiscal_year
    
    # ใช้ logic เดียวกันกับ receipt_report_view สำหรับ filter
    if request.user.has_permission('receipt_view_all'):
        receipts = Receipt.objects.all()
        view_scope = "ทุกหน่วยงาน"
    else:
        receipts = Receipt.objects.filter(department__name=request.user.department)
        view_scope = f"หน่วยงาน: {request.user.department}"
    
    # รับค่า filter
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    department_filter = request.GET.get('department')
    status_filter = request.GET.get('status')
    search_query = request.GET.get('q', '').strip()
    
    # Apply filters (copy logic from receipt_report_view)
    if date_from:
        try:
            receipts = receipts.filter(receipt_date__gte=datetime.strptime(date_from, '%Y-%m-%d').date())
        except ValueError:
            pass
    
    if date_to:
        try:
            receipts = receipts.filter(receipt_date__lte=datetime.strptime(date_to, '%Y-%m-%d').date())
        except ValueError:
            pass
    
    if department_filter and request.user.has_permission('receipt_view_all'):
        receipts = receipts.filter(department__name=department_filter)
    
    if status_filter:
        receipts = receipts.filter(status=status_filter)
    
    if search_query:
        receipts = receipts.filter(
            Q(receipt_number__icontains=search_query) |
            Q(recipient_name__icontains=search_query)
        )
    
    # เรียงลำดับและ prefetch
    receipts = receipts.select_related('department', 'created_by').prefetch_related('items').order_by('-created_at')
    
    # สร้าง Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "รายงานใบสำคัญรับเงิน"
    
    # กำหนดสี
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header information
    current_fiscal = get_current_fiscal_year()
    ws['A1'] = "รายงานใบสำคัญรับเงิน"
    ws['A2'] = f"ประจำปีงบประมาณ {current_fiscal}"
    
    # แสดงหน่วยงาน
    if department_filter:
        ws['A3'] = f"ชื่อหน่วยงาน: {department_filter}"
    else:
        ws['A3'] = f"ขอบเขต: {view_scope}"
    
    # แสดงช่วงวันที่
    date_range = ""
    if date_from and date_to:
        date_range = f"ระหว่างวันที่ {date_from} ถึง {date_to}"
    elif date_from:
        date_range = f"ตั้งแต่วันที่ {date_from}"
    elif date_to:
        date_range = f"จนถึงวันที่ {date_to}"
    else:
        date_range = "ทุกช่วงเวลา"
    
    ws['A4'] = date_range
    
    # ปรับ style สำหรับ header
    for row in range(1, 5):
        ws[f'A{row}'].font = Font(bold=True, size=14 if row == 1 else 12)
        ws[f'A{row}'].alignment = Alignment(horizontal='center' if row == 1 else 'left')
    
    # Merge cells สำหรับ title
    ws.merge_cells('A1:H1')
    
    # Table headers (row 6)
    headers = [
        'ลำดับ',
        'ใบสำคัญเลขที่', 
        'วันที่ขอ',
        'รายการ',
        'จำนวนเงิน',
        'ผู้รับเงิน',
        'ผู้จ่ายเงิน',
        'หมายเหตุ'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    row_num = 7
    total_amount = 0
    
    for index, receipt in enumerate(receipts, 1):
        # รวมรายการสินค้า
        items_text = []
        for item in receipt.items.all():
            items_text.append(f"{item.description} ({item.amount:,.2f} บาท)")
        items_display = "; ".join(items_text) if items_text else "-"
        
        # ผู้จ่ายเงิน (created_by)
        payer = receipt.created_by.get_display_name() if receipt.created_by else "-"
        
        # สถานะเป็นหมายเหตุ
        status_map = {
            'draft': 'ร่าง',
            'completed': 'เสร็จสิ้น', 
            'cancelled': 'ยกเลิก'
        }
        notes = status_map.get(receipt.status, receipt.status)
        
        # เพิ่มข้อมูลลงในแถว
        row_data = [
            index,  # ลำดับ
            receipt.receipt_number,  # ใบสำคัญเลขที่
            receipt.receipt_date.strftime('%d/%m/%Y') if receipt.receipt_date else '-',  # วันที่ขอ
            items_display,  # รายการ
            receipt.total_amount,  # จำนวนเงิน
            receipt.recipient_name,  # ผู้รับเงิน
            payer,  # ผู้จ่ายเงิน
            notes  # หมายเหตุ
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col)
            cell.value = value
            cell.border = border
            
            # จัดตำแหน่ง
            if col in [1, 5]:  # ลำดับ, จำนวนเงิน
                cell.alignment = Alignment(horizontal='right')
            elif col == 3:  # วันที่
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='left')
            
            # Format ตัวเลข
            if col == 5:  # จำนวนเงิน
                cell.number_format = '#,##0.00'
        
        if receipt.status == 'completed':
            total_amount += receipt.total_amount
        
        row_num += 1
    
    # แถวสรุปยอดรวม
    summary_row = row_num + 1
    ws[f'D{summary_row}'] = "รวมทั้งสิ้น"
    ws[f'E{summary_row}'] = total_amount
    
    # Style สำหรับแถวสรุป
    for col in ['D', 'E']:
        cell = ws[f'{col}{summary_row}']
        cell.font = Font(bold=True)
        cell.border = border
        if col == 'E':
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right')
        else:
            cell.alignment = Alignment(horizontal='center')
    
    # ปรับขนาดคอลัมน์
    column_widths = [8, 18, 12, 40, 15, 25, 20, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # สร้าง response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # สร้างชื่อไฟล์
    filename = f"รายงานใบสำคัญรับเงิน_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # บันทึก workbook
    wb.save(response)
    
    return response


# Template Management Views (Admin Only)
@login_required
def receipt_templates_list(request):
    """แสดงรายการ Template ทั้งหมด (Admin เท่านั้น)"""
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, 'คุณไม่มีสิทธิ์เข้าถึงหน้านี้')
        return redirect('dashboard')
    
    templates = ReceiptTemplate.objects.all().order_by('category', 'name')
    
    context = {
        'title': 'จัดการรายการสำเร็จรูป',
        'templates': templates,
    }
    
    return render(request, 'accounts/receipt_templates_list.html', context)


@login_required
def receipt_template_create(request):
    """สร้าง Template ใหม่ (Admin เท่านั้น)"""
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, 'คุณไม่มีสิทธิ์เข้าถึงหน้านี้')
        return redirect('dashboard')
    
    if request.method == 'POST':
        # Process form data
        name = request.POST.get('name', '').strip()
        category = request.POST.get('category', '').strip()
        max_amount = request.POST.get('max_amount', '').strip()
        fixed_amount = request.POST.get('fixed_amount', '').strip()
        is_active = request.POST.get('is_active') == 'on'
        
        # Validation
        if not name:
            messages.error(request, 'กรุณากรอกชื่อรายการ')
        elif ReceiptTemplate.objects.filter(name=name).exists():
            messages.error(request, 'ชื่อรายการนี้มีอยู่แล้ว')
        else:
            # Create template
            template = ReceiptTemplate.objects.create(
                name=name,
                category=category,
                max_amount=float(max_amount) if max_amount else None,
                fixed_amount=float(fixed_amount) if fixed_amount else None,
                is_active=is_active
            )
            
            messages.success(request, f'สร้างรายการ "{template.name}" เรียบร้อยแล้ว')
            return redirect('receipt_templates_list')
    
    context = {
        'title': 'เพิ่มรายการสำเร็จรูป',
    }
    
    return render(request, 'accounts/receipt_template_create.html', context)


@login_required
def receipt_template_edit(request, template_id):
    """แก้ไข Template (Admin เท่านั้น)"""
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, 'คุณไม่มีสิทธิ์เข้าถึงหน้านี้')
        return redirect('dashboard')
    
    template = get_object_or_404(ReceiptTemplate, id=template_id)
    
    if request.method == 'POST':
        # Process form data
        name = request.POST.get('name', '').strip()
        category = request.POST.get('category', '').strip()
        max_amount = request.POST.get('max_amount', '').strip()
        fixed_amount = request.POST.get('fixed_amount', '').strip()
        is_active = request.POST.get('is_active') == 'on'
        
        # Validation
        if not name:
            messages.error(request, 'กรุณากรอกชื่อรายการ')
        elif ReceiptTemplate.objects.filter(name=name).exclude(id=template.id).exists():
            messages.error(request, 'ชื่อรายการนี้มีอยู่แล้ว')
        else:
            # Update template
            template.name = name
            template.category = category
            template.max_amount = float(max_amount) if max_amount else None
            template.fixed_amount = float(fixed_amount) if fixed_amount else None
            template.is_active = is_active
            template.save()
            
            messages.success(request, f'อัปเดตรายการ "{template.name}" เรียบร้อยแล้ว')
            return redirect('receipt_templates_list')
    
    context = {
        'title': 'แก้ไขรายการสำเร็จรูป',
        'template': template,
    }
    
    return render(request, 'accounts/receipt_template_edit.html', context)


@login_required
def receipt_template_delete(request, template_id):
    """ลบ Template (Admin เท่านั้น)"""
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, 'คุณไม่มีสิทธิ์เข้าถึงหน้านี้')
        return redirect('dashboard')
    
    template = get_object_or_404(ReceiptTemplate, id=template_id)
    
    if request.method == 'POST':
        template_name = template.name
        template.delete()
        messages.success(request, f'ลบรายการ "{template_name}" เรียบร้อยแล้ว')
    
    return redirect('receipt_templates_list')


@login_required
def receipt_edit_view(request, receipt_id):
    """
    หน้าแก้ไขใบสำคัญรับเงิน (เฉพาะร่าง)
    """
    try:
        receipt = Receipt.objects.get(id=receipt_id)
    except Receipt.DoesNotExist:
        messages.error(request, 'ไม่พบใบสำคัญรับเงินที่ต้องการ')
        return redirect('receipt_list')
    
    # ตรวจสอบสิทธิ์
    if not request.user.has_permission('receipt_create'):
        messages.error(request, 'คุณไม่มีสิทธิ์แก้ไขใบสำคัญรับเงิน')
        return redirect('receipt_list')
    
    # ตรวจสอบว่าเป็นร่างและเป็นของผู้ใช้เท่านั้น
    if receipt.status != 'draft':
        messages.error(request, 'สามารถแก้ไขได้เฉพาะใบสำคัญสถานะ "ร่าง" เท่านั้น')
        return redirect('receipt_detail', receipt_id=receipt_id)
    
    if receipt.created_by != request.user:
        messages.error(request, 'คุณสามารถแก้ไขได้เฉพาะใบสำคัญของตัวเองเท่านั้น')
        return redirect('receipt_detail', receipt_id=receipt_id)
    
    # ดึงข้อมูลที่จำเป็น
    try:
        user_department = Department.objects.get(name=request.user.department)
    except Department.DoesNotExist:
        messages.error(request, 'ไม่พบหน่วยงานของคุณในระบบ')
        return redirect('receipt_detail', receipt_id=receipt_id)
    
    # ดึงรายการสำเร็จรูป
    receipt_templates = ReceiptTemplate.objects.filter(is_active=True).order_by('category', 'name')
    
    # ดึงรายการใบสำคัญ
    receipt_items = receipt.items.all().order_by('order')
    
    context = {
        'title': f'แก้ไขใบสำคัญรับเงิน - {receipt.receipt_number}',
        'receipt': receipt,
        'receipt_items': receipt_items,
        'department': user_department,
        'receipt_templates': receipt_templates,
        'user_full_name': request.user.get_display_name(),
    }
    
    return render(request, 'accounts/receipt_edit.html', context)


@login_required  
def receipt_update_ajax(request, receipt_id):
    """
    อัปเดตใบสำคัญรับเงิน (เฉพาะร่าง) ผ่าน AJAX
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)
    
    try:
        receipt = Receipt.objects.get(id=receipt_id)
    except Receipt.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'ไม่พบใบสำคัญรับเงิน'}, status=404)
    
    # ตรวจสอบสิทธิ์
    if not request.user.has_permission('receipt_create'):
        return JsonResponse({'success': False, 'message': 'ไม่มีสิทธิ์แก้ไขใบสำคัญรับเงิน'}, status=403)
    
    # ตรวจสอบว่าเป็นร่างและเป็นของผู้ใช้
    if receipt.status != 'draft':
        return JsonResponse({'success': False, 'message': 'สามารถแก้ไขได้เฉพาะใบสำคัญสถานะ "ร่าง" เท่านั้น'}, status=400)
    
    if receipt.created_by != request.user:
        return JsonResponse({'success': False, 'message': 'คุณสามารถแก้ไขได้เฉพาะใบสำคัญของตัวเองเท่านั้น'}, status=403)
    
    try:
        data = json.loads(request.body)
        
        # ตรวจสอบข้อมูลจำเป็น
        required_fields = ['recipient_name', 'recipient_address', 'recipient_id_card']
        for field in required_fields:
            if not data.get(field):
                return JsonResponse({'success': False, 'message': f'กรุณากรอก{field}'}, status=400)
        
        # ตรวจสอบจำนวนเงิน
        try:
            total_amount = float(data.get('total_amount', 0))
            if total_amount <= 0:
                return JsonResponse({'success': False, 'message': 'จำนวนเงินต้องมากกว่า 0'}, status=400)
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'message': 'จำนวนเงินไม่ถูกต้อง'}, status=400)
        
        # ตรวจสอบและอัปเดต receipt_date
        from datetime import datetime
        from utils.fiscal_year import get_current_fiscal_year, get_fiscal_year_dates

        if data.get('receipt_date'):
            try:
                # Parse วันที่จาก string (YYYY-MM-DD)
                new_receipt_date = datetime.strptime(data.get('receipt_date'), '%Y-%m-%d').date()

                # ตรวจสอบวันที่
                today = datetime.now().date()

                # ห้ามวันที่อนาคต
                if new_receipt_date > today:
                    return JsonResponse({'success': False, 'message': 'ไม่สามารถกำหนดวันที่ล่วงหน้าได้'}, status=400)

                # ตรวจสอบว่าไม่เกินต้นปีงบประมาณ
                current_fy = get_current_fiscal_year()
                fiscal_start, fiscal_end = get_fiscal_year_dates(current_fy)

                if new_receipt_date < fiscal_start:
                    return JsonResponse({
                        'success': False,
                        'message': f'วันที่ย้อนหลังเกินกำหนด (ย้อนหลังได้ถึง {fiscal_start.strftime("%d/%m/%Y")} เท่านั้น)'
                    }, status=400)

                receipt.receipt_date = new_receipt_date

            except ValueError:
                return JsonResponse({'success': False, 'message': 'รูปแบบวันที่ไม่ถูกต้อง'}, status=400)

        new_status = data.get('status', 'draft')

        # ถ้าเปลี่ยนจาก draft เป็น completed และยังไม่มี receipt_date ให้ตั้งวันที่ปัจจุบัน
        if new_status == 'completed' and not receipt.receipt_date:
            receipt.receipt_date = datetime.now().date()

        # อัปเดตข้อมูลใบสำคัญ
        receipt.recipient_name = data.get('recipient_name', '')
        receipt.recipient_address = data.get('recipient_address', '')
        receipt.recipient_postal_code = data.get('recipient_postal_code', '')
        receipt.recipient_id_card = data.get('recipient_id_card', '')
        receipt.is_loan = data.get('is_loan', False)
        receipt.status = new_status
        
        # ลบรายการเก่าทั้งหมด
        receipt.items.all().delete()
        
        # สร้างรายการใหม่
        items_data = data.get('items', [])
        calculated_total = 0
        
        for idx, item_data in enumerate(items_data, 1):
            try:
                amount = float(item_data.get('amount', 0))
                if amount <= 0:
                    continue
                calculated_total += amount
                
                # จัดการ template_id
                template_id = item_data.get('template_id')
                if template_id and str(template_id).lower() in ['null', 'none', '']:
                    template_id = None
                elif template_id:
                    try:
                        template_id = int(template_id)
                    except (ValueError, TypeError):
                        template_id = None
                
                ReceiptItem.objects.create(
                    receipt=receipt,
                    template_id=template_id,
                    description=item_data.get('description', ''),
                    amount=amount,
                    order=idx
                )
            except (ValueError, TypeError) as e:
                return JsonResponse({'success': False, 'message': f'รายการที่ {idx} มีข้อผิดพลาด: {str(e)}'}, status=400)
        
        # อัปเดต total_amount และ force regenerate total_amount_text
        receipt.total_amount = calculated_total
        receipt.total_amount_text = Receipt.convert_amount_to_thai_text(calculated_total)
        receipt.save()  # จะ auto-generate ทั้ง hash และ qr code ใหม่

        # บันทึก Change Log
        ReceiptChangeLog.log_change(
            receipt=receipt,
            action='updated',
            user=request.user,
            notes=f'แก้ไขใบสำคัญรับเงิน (สถานะ: {receipt.get_status_display()})'
        )

        return JsonResponse({
            'success': True,
            'message': 'อัปเดตใบสำคัญรับเงินเรียบร้อย',
            'receipt_id': receipt.id,
            'receipt_number': receipt.receipt_number,
            'receipt_date': receipt.receipt_date.strftime('%d/%m/%Y') if receipt.receipt_date else '',
            'verification_url': receipt.get_verification_url()
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'ข้อมูล JSON ไม่ถูกต้อง'}, status=400)
    except Exception as e:
        import logging
        logging.error(f'Receipt update error: {str(e)}')
        return JsonResponse({'success': False, 'message': f'เกิดข้อผิดพลาด: {str(e)}'}, status=500)


@login_required
def receipt_complete_draft_ajax(request, receipt_id):
    """
    เปลี่ยนสถานะใบสำคัญจาก draft เป็น completed
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)

    try:
        receipt = Receipt.objects.get(id=receipt_id)
    except Receipt.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'ไม่พบใบสำคัญรับเงิน'}, status=404)

    # ตรวจสอบสิทธิ์
    if receipt.created_by != request.user:
        return JsonResponse({'success': False, 'message': 'คุณสามารถบันทึกได้เฉพาะใบสำคัญของตัวเองเท่านั้น'}, status=403)

    # ตรวจสอบสถานะ
    if receipt.status != 'draft':
        return JsonResponse({'success': False, 'message': 'สามารถบันทึกได้เฉพาะใบสำคัญสถานะ "ร่าง" เท่านั้น'}, status=400)

    # เปลี่ยนสถานะเป็น completed
    from datetime import datetime
    old_status = receipt.status
    receipt.status = 'completed'

    # ตั้งค่า receipt_date เฉพาะกรณีที่ยังไม่มีการกำหนดวันที่ (เผื่อกรณี legacy data)
    # ไม่แทนที่วันที่ที่ผู้ใช้กำหนดไว้แล้ว
    if not receipt.receipt_date:
        receipt.receipt_date = datetime.now().date()

    receipt.save()

    # บันทึก Change Log
    ReceiptChangeLog.log_change(
        receipt=receipt,
        action='updated',
        user=request.user,
        field_name='status',
        old_value=old_status,
        new_value='completed',
        notes='เปลี่ยนสถานะจากร่างเป็นเสร็จสิ้น'
    )

    return JsonResponse({
        'success': True,
        'message': f'บันทึกใบสำคัญ {receipt.receipt_number} เรียบร้อยแล้ว',
        'receipt_id': receipt.id,
        'receipt_number': receipt.receipt_number
    })


@login_required
def cancel_request_list_view(request):
    """
    รายการคำร้องขอยกเลิกใบสำคัญรับเงิน
    - Basic User: ดูเฉพาะคำร้องของตัวเอง
    - Department Manager: ดูคำร้องของหน่วยงานตัวเอง + ที่สามารถอนุมัติได้
    - Admin: ดูทั้งหมด
    """
    from .models import ReceiptCancelRequest
    
    # กรองข้อมูลตามสิทธิ์
    if request.user.has_permission('receipt_view_all'):
        # Admin: ดูทั้งหมด
        cancel_requests = ReceiptCancelRequest.objects.all()
        view_scope = 'ทั้งหมด'
    elif request.user.has_permission('receipt_cancel_approve_manager'):
        # Senior Manager: ดูทั้งหมด
        cancel_requests = ReceiptCancelRequest.objects.all()
        view_scope = 'ทั้งหมด (Senior Manager)'
    elif request.user.has_permission('receipt_cancel_approve'):
        # Department Manager: ดูเฉพาะหน่วยงานตัวเอง
        cancel_requests = ReceiptCancelRequest.objects.filter(
            receipt__department__name=request.user.department
        )
        view_scope = f'หน่วยงาน: {request.user.department}'
    else:
        # Basic User: ดูเฉพาะของตัวเอง
        cancel_requests = ReceiptCancelRequest.objects.filter(
            requested_by=request.user
        )
        view_scope = 'คำขอของฉัน'
    
    # คำนวณสถิติก่อน filter (เพื่อให้เห็นภาพรวมทั้งหมด)
    base_queryset = cancel_requests
    stats = {
        'total': base_queryset.count(),
        'pending': base_queryset.filter(status='pending').count(),
        'approved': base_queryset.filter(status='applied').count(),  # ดำเนินการแล้ว
        'rejected': base_queryset.filter(status='rejected').count(),
    }

    # Filter และ Search
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('q', '')

    if status_filter:
        cancel_requests = cancel_requests.filter(status=status_filter)

    if search_query:
        cancel_requests = cancel_requests.filter(
            Q(request_number__icontains=search_query) |
            Q(receipt__receipt_number__icontains=search_query) |
            Q(receipt__recipient_name__icontains=search_query) |
            Q(cancel_reason__icontains=search_query)
        )

    # เพิ่มการเช็คสิทธิ์อนุมัติสำหรับแต่ละ request
    for cancel_request in cancel_requests:
        cancel_request.can_be_approved_by = cancel_request.can_be_approved_by(request.user)

    # Pagination
    paginator = Paginator(cancel_requests.select_related('receipt', 'requested_by', 'approved_by').order_by('-created_at'), 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # สถิติเดิม (เก็บไว้เพื่อ backward compatibility)
    pending_count = stats['pending']

    context = {
        'title': 'รายการคำขออนุมัติยกเลิก',
        'cancel_requests': page_obj,
        'view_scope': view_scope,
        'status_filter': status_filter,
        'search_query': search_query,
        'pending_count': pending_count,
        'stats': stats,
    }

    return render(request, 'accounts/cancel_request_list.html', context)


@login_required
def audit_log_view(request):
    """
    หน้าแสดงประวัติการเปลี่ยนแปลงทั้งหมด (Audit Log)
    สำหรับ Admin และผู้มีสิทธิ์
    """
    from django.core.paginator import Paginator
    from django.db.models import Q

    # ตรวจสอบสิทธิ์ (อนุญาตเฉพาะ Admin หรือ Department Manager)
    if not (request.user.has_permission('receipt_view_all') or request.user.is_staff):
        messages.error(request, 'คุณไม่มีสิทธิ์เข้าถึงหน้านี้')
        return redirect('dashboard')

    # Get filter parameters
    action_filter = request.GET.get('action', '')
    department_filter = request.GET.get('department', '')
    user_filter = request.GET.get('user', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search_query = request.GET.get('q', '')

    # Base queryset
    logs = ReceiptChangeLog.objects.select_related('receipt', 'user', 'edit_request').all()

    # Apply filters
    if action_filter:
        logs = logs.filter(action=action_filter)

    if department_filter:
        logs = logs.filter(receipt__department__id=department_filter)

    if user_filter:
        logs = logs.filter(user__id=user_filter)

    if date_from:
        from datetime import datetime
        logs = logs.filter(created_at__gte=datetime.strptime(date_from, '%Y-%m-%d'))

    if date_to:
        from datetime import datetime
        logs = logs.filter(created_at__lte=datetime.strptime(date_to, '%Y-%m-%d').replace(hour=23, minute=59, second=59))

    if search_query:
        logs = logs.filter(
            Q(receipt__receipt_number__icontains=search_query) |
            Q(user__full_name__icontains=search_query) |
            Q(notes__icontains=search_query)
        )

    # Pagination
    paginator = Paginator(logs.order_by('-created_at'), 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Get filter options
    departments = Department.objects.filter(is_active=True).order_by('name')
    actions = ReceiptChangeLog.ACTION_CHOICES

    # Get users who have made changes
    from django.db.models import Count
    active_users = User.objects.filter(
        id__in=ReceiptChangeLog.objects.values_list('user_id', flat=True).distinct()
    ).order_by('full_name')

    context = {
        'title': 'ประวัติการเปลี่ยนแปลง (Audit Log)',
        'logs': page_obj,
        'departments': departments,
        'actions': actions,
        'active_users': active_users,
        'action_filter': action_filter,
        'department_filter': department_filter,
        'user_filter': user_filter,
        'date_from': date_from,
        'date_to': date_to,
        'search_query': search_query,
        'total_count': logs.count(),
    }

    return render(request, 'accounts/audit_log.html', context)


@login_required
def audit_log_excel_export(request):
    """
    Export Audit Log เป็น Excel
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    from datetime import datetime

    # ตรวจสอบสิทธิ์
    if not (request.user.has_permission('receipt_view_all') or request.user.is_staff):
        messages.error(request, 'คุณไม่มีสิทธิ์ export ข้อมูล')
        return redirect('audit_log')

    # Get same filters as main view
    action_filter = request.GET.get('action', '')
    department_filter = request.GET.get('department', '')
    user_filter = request.GET.get('user', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search_query = request.GET.get('q', '')

    # Build queryset with same filters
    logs = ReceiptChangeLog.objects.select_related('receipt', 'user', 'edit_request').all()

    if action_filter:
        logs = logs.filter(action=action_filter)
    if department_filter:
        logs = logs.filter(receipt__department__id=department_filter)
    if user_filter:
        logs = logs.filter(user__id=user_filter)
    if date_from:
        logs = logs.filter(created_at__gte=datetime.strptime(date_from, '%Y-%m-%d'))
    if date_to:
        logs = logs.filter(created_at__lte=datetime.strptime(date_to, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
    if search_query:
        from django.db.models import Q
        logs = logs.filter(
            Q(receipt__receipt_number__icontains=search_query) |
            Q(user__full_name__icontains=search_query) |
            Q(notes__icontains=search_query)
        )

    logs = logs.order_by('-created_at')

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Audit Log'

    # Headers
    headers = ['วันที่-เวลา', 'เลขที่ใบสำคัญ', 'การดำเนินการ', 'ฟิลด์', 'ค่าเดิม', 'ค่าใหม่', 'ผู้ดำเนินการ', 'หมายเหตุ']
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Add data
    for log in logs:
        ws.append([
            log.created_at.strftime('%d/%m/%Y %H:%M:%S'),
            log.receipt.receipt_number,
            log.get_action_display(),
            log.field_name or '-',
            log.old_value or '-',
            log.new_value or '-',
            log.user.get_display_name() if log.user else '-',
            log.notes or '-'
        ])

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 40

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'audit_log_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response


# ===== USER ACTIVITY LOG (ADMIN ONLY) =====

@login_required
def user_activity_log_view(request):
    """
    หน้าแสดงประวัติการใช้งานระบบ (User Activity Log)
    สำหรับผู้ดูแลระบบเท่านั้น
    """
    # Permission check - Admin only
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, 'คุณไม่มีสิทธิ์เข้าถึงหน้านี้ (เฉพาะผู้ดูแลระบบ)')
        return redirect('dashboard')

    # Get all activity logs
    logs = UserActivityLog.objects.select_related('user').all()

    # Filter by action
    action_filter = request.GET.get('action', '')
    if action_filter:
        logs = logs.filter(action=action_filter)

    # Filter by user
    user_filter = request.GET.get('user', '')
    if user_filter:
        logs = logs.filter(user_id=user_filter)

    # Filter by date range
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if date_from:
        from datetime import datetime
        date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        logs = logs.filter(created_at__date__gte=date_from_obj)
    if date_to:
        from datetime import datetime
        date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
        logs = logs.filter(created_at__date__lte=date_to_obj)

    # Search in username, IP address
    search_query = request.GET.get('q', '')
    if search_query:
        logs = logs.filter(
            Q(username_attempted__icontains=search_query) |
            Q(ip_address__icontains=search_query) |
            Q(user__full_name__icontains=search_query) |
            Q(notes__icontains=search_query)
        )

    # Count total before pagination
    total_count = logs.count()

    # Pagination
    paginator = Paginator(logs.order_by('-created_at'), 50)  # 50 items per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Get filter options
    actions = UserActivityLog.ACTION_CHOICES
    active_users = User.objects.filter(is_active=True).order_by('full_name')

    context = {
        'title': 'ประวัติการใช้งานระบบ',
        'logs': page_obj,
        'total_count': total_count,
        'actions': actions,
        'active_users': active_users,
        'action_filter': action_filter,
        'user_filter': user_filter,
        'date_from': date_from,
        'date_to': date_to,
        'search_query': search_query,
    }

    return render(request, 'accounts/user_activity_log.html', context)


@login_required
def user_activity_log_excel_export(request):
    """Export User Activity Log เป็น Excel"""
    # Permission check - Admin only
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, 'คุณไม่มีสิทธิ์ใช้งานฟังก์ชันนี้')
        return redirect('dashboard')

    # Get filtered logs (same logic as view)
    logs = UserActivityLog.objects.select_related('user').all()

    # Apply same filters as view
    action_filter = request.GET.get('action', '')
    if action_filter:
        logs = logs.filter(action=action_filter)

    user_filter = request.GET.get('user', '')
    if user_filter:
        logs = logs.filter(user_id=user_filter)

    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if date_from:
        from datetime import datetime
        date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        logs = logs.filter(created_at__date__gte=date_from_obj)
    if date_to:
        from datetime import datetime
        date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
        logs = logs.filter(created_at__date__lte=date_to_obj)

    search_query = request.GET.get('q', '')
    if search_query:
        logs = logs.filter(
            Q(username_attempted__icontains=search_query) |
            Q(ip_address__icontains=search_query) |
            Q(user__full_name__icontains=search_query) |
            Q(notes__icontains=search_query)
        )

    logs = logs.order_by('-created_at')

    # Create Excel workbook
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'User Activity Log'

    # Headers
    headers = ['วันที่-เวลา', 'ผู้ใช้', 'การดำเนินการ', 'IP Address', 'User Agent', 'หมายเหตุ']
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add data
    for log in logs:
        ws.append([
            log.created_at.strftime('%d/%m/%Y %H:%M:%S'),
            log.user.get_display_name() if log.user else log.username_attempted,
            log.get_action_display(),
            log.ip_address or '-',
            (log.user_agent[:50] + '...') if len(log.user_agent) > 50 else log.user_agent or '-',
            log.notes or '-'
        ])

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 40

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'user_activity_log_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response